"""
Excel annexure writer matching Athena PowerTech's deliverable format.

Two formats:

**Format `E_F`** — the modern HMEL-style deliverable. Two sheets:

  * **Annexure E** — "Run to Run Comparison" of every run-2 feature
    against its run-1 partner (or the 10 %-WT assumption when unmatched).
    Three-row header: title (merged), group headers (merged in 5 spans),
    field headers; CGR (mm/yr) is the rightmost column.

  * **Annexure F** — "Metal Loss Anomalies" — the per-feature catalogue
    with location, geometry, surface, and predicted repair date. Three
    title rows + one header row.

**Format `B_C_D`** — the older GAIL Samakhiali-style deliverable. Three
sheets:

  * **Annexure B** — matched defects with run-to-run depths and CGR
  * **Annexure C** — ALL defects assessed via B31G (current + 10-year
    forward projection of depth, SOP, ERF)
  * **Annexure D** — circumferential defects only, assessed via Kastner

Both formats place an optional "QA Issues" sheet at the end when a
`FlagReport` is supplied.

Cell styling follows what the reference files use:
  * Title row: bold 12pt, fill `FFC000` (yellow/orange), centered + wrapped
  * Sub-title rows: bold 11pt
  * Group/field header rows: bold 10pt, fill `BDD7EE` (light blue),
    centered, thin black borders
  * Data rows: regular 10pt, thin black borders
  * Numeric formats: distance `0.000`, depth `0.00`, WT `0.0`, CGR `0.0000`
  * Orientation rendered as `hh:mm:ss` text (matches the published files)
  * Repair date: `dd-mm-yyyy` when triggered, "After {horizon end}" string
    when the feature stays below both triggers for the whole horizon
  * All rows sorted ascending by run-2 absolute distance

The structures (merged ranges, column widths, header text) were read
off the published reference workbooks in `/examples/` — see
`docs/REPORT_FORMATS.md` for the audit table.
"""
from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Iterable

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.core.cgr import CGRResult
from src.core.repair_predictor import (
    TRIGGER_DEPTH_80,
    TRIGGER_ERF_1,
    TRIGGER_NONE,
    horizon_end_date,
)
from src.models import FFPMethod, FFPResult, Feature, Pipeline, Project, Surface


# ---------------------------------------------------------------------------
# Styling constants (matched to the reference workbooks)
# ---------------------------------------------------------------------------

_THIN = Side(style="thin", color="000000")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_TITLE_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
_HEADER_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

_TITLE_FONT = Font(bold=True, size=12)
_SUBTITLE_FONT = Font(bold=True, size=11)
_HEADER_FONT = Font(bold=True, size=10)
_DATA_FONT = Font(size=10)

_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


# ---------------------------------------------------------------------------
# Annexure E/F layouts (1-indexed columns)
# ---------------------------------------------------------------------------

# Annexure E columns A..N (14 from v0.2.6 — added CGR raw before CGR)
#
# v0.2.6: Annexure B / results_ili_comparison gains a "CGR raw (mm/yr)"
# column (column 13, M) showing the per-feature pre-floor rate. The
# existing "CGR (mm/yr)" column moves to column 14 (N) and continues to
# hold the post-floor value (what FFP projection consumes).
#
# In feature_specific / population_only CGR modes the two columns are
# identical; in hybrid mode they may differ. Engineers comparing the
# tool's output against hand-calculated published-report CGRs need the
# pre-floor value to match the raw formula `(d_new − d_unmatched)/Δt`.
_ANNEX_E_COLUMNS = [
    ("S.N.",               12, "0"),
    ("Anomaly ID",         15, "@"),
    ("Wall Thickness, (mm)", 14, "0.0"),
    ("Joint Number",       14, "0"),
    ("Abs Dist new",       14, "0.000"),
    ("Abs Dist old",       14, "0.000"),
    ("Depth new",          12, "0.00"),
    ("Depth old",          12, "0.00"),
    ("Orient new",         12, "@"),
    ("Orient old",         12, "@"),
    ("Surface new",        11, "@"),
    ("Surface old",        11, "@"),
    ("CGR raw (mm/yr)",    14, "0.0000"),   # v0.2.6 new — pre-floor
    ("CGR (mm/yr)",        14, "0.0000"),   # post-floor (what FFP uses)
]

# Annexure F columns A..P (16)
_ANNEX_F_COLUMNS = [
    ("S.N.",                              8,  "0"),
    ("Feature ID",                        12, "@"),
    ("Absolute Distance\n[m]",            16, "0.000"),
    ("Latitude",                          16, "0.0000000000"),
    ("Longitude",                         16, "0.0000000000"),
    ("Joint No.",                         10, "0"),
    ("Joint Length (m)",                  14, "0.000"),
    ("Distance to closest weld (m)",      16, "0.000"),
    ("Event",                             24, "@"),
    ("Surface",                           10, "@"),
    ("Wall Thickness [mm]",               14, "0.0"),
    ("Orientation (hh:mm)",               14, "@"),
    ("Reported Depth\n[% WT]",            14, "0.00"),
    ("Length (mm)",                       12, "0"),
    ("Width (mm)",                        12, "0"),
    ("Predicted Repair year-\nEffective Repair Date", 24, "@"),
]


# ---------------------------------------------------------------------------
# AnnexureWriter
# ---------------------------------------------------------------------------

class AnnexureWriter:
    """Generate the Excel deliverable in either the Annexure E/F (modern
    HMEL) or Annexure B/C/D (older GAIL) format.
    """

    def __init__(self, *, horizon_years: int = 10):
        self.horizon_years = horizon_years

    # ------------------------------------------------------------------

    def write(
        self,
        *,
        cgr_results: Iterable[CGRResult] = (),
        ffp_results: Iterable[FFPResult] = (),
        repair_predictions: Iterable = (),
        flag_report: Any = None,
        project: Project | None = None,
        pipeline: Pipeline | None = None,
        output_path: str | Path,
        format: str | None = None,
        topics: list[tuple[str, str]] | None = None,
        years_between: float | None = None,
        match_result: Any = None,            # accepted for API compat, unused
    ) -> None:
        """Build the annexure XLSX.

        v0.2.5: two dispatch modes:

          * **Topic mode** (preferred, new) — caller passes
            ``topics=[(topic_id, letter), ...]``. Each entry becomes
            one sheet, ordered as given, named per the engineer's
            chosen letter. Topics are looked up in
            :data:`src.reports.topic_registry.TOPIC_REGISTRY`. Unknown
            IDs are silently skipped (the YAML parser raises on them
            up-front, so anything reaching this point is trusted).

          * **Legacy preset mode** (fallback for old callers) —
            ``format="E_F"`` or ``"B_C_D"``. Same behaviour as
            v0.2.0–v0.2.4. Used when ``topics is None``.

        If both arguments are missing, defaults to ``format="E_F"`` so
        ad-hoc test code without a topic list keeps working.
        """
        out_path = Path(output_path)
        out_path.parent.mkdir(parents=True, exist_ok=True)

        # Resolve pipeline from project if not given.
        if pipeline is None and project is not None:
            pipeline = project.pipeline

        wb = Workbook()
        wb.remove(wb.active)            # drop the default empty sheet

        # ---- v0.2.5 topic-walker mode --------------------------------
        if topics is not None:
            from types import SimpleNamespace
            from src.reports.topic_registry import TOPIC_REGISTRY
            run2_year = (
                project.run_2.inspection_date.year
                if (project and project.run_2
                    and project.run_2.inspection_date) else None
            )
            results_ns = SimpleNamespace(
                cgr_results=list(cgr_results),
                ffp_results=list(ffp_results),
                repair_predictions=list(repair_predictions),
                flag_report=flag_report,
                years_between=(
                    years_between
                    if years_between is not None
                    else (project.years_between_runs if project else 0.0)
                ),
            )
            for topic_id, letter in topics:
                topic = TOPIC_REGISTRY.get(topic_id)
                if topic is None:
                    # YAML parser is responsible for raising on unknown
                    # IDs; if one slipped through, skip rather than
                    # crashing the whole report.
                    continue
                sheet_name, title_text = make_topic_sheet_name(
                    letter, topic.display_name,
                )
                topic.writer(
                    wb, sheet_name, project, results_ns, run2_year,
                    title_text=title_text,
                )
            wb.save(out_path)
            return

        # ---- Legacy preset mode (E_F / B_C_D) ------------------------
        # Map FFPResult / RepairPrediction by feature_id for lookup.
        ffp_by_id = {r.feature_id: r for r in ffp_results}
        pred_by_id = {p.feature_id: p for p in repair_predictions}

        # Year labels for Annexure E group headers.
        year_new, year_old = _resolve_year_labels(project)
        section_name, pipeline_name = _resolve_section_labels(project, pipeline)

        # Sort CGR results by run-2 absolute distance ascending.
        cgrs_sorted = sorted(
            cgr_results,
            key=lambda r: (r.feature.abs_distance_m or 0.0, str(r.feature.anomaly_id)),
        )

        fmt = (format or "E_F").upper().replace("/", "_")
        if fmt in ("E_F", "EF"):
            self._write_annexure_e(
                wb, cgrs_sorted, year_new=year_new, year_old=year_old,
                section_name=section_name,
            )
            self._write_annexure_f(
                wb, cgrs_sorted, ffp_by_id, pred_by_id, pipeline=pipeline,
                section_name=section_name, pipeline_name=pipeline_name,
                project=project,
            )
        elif fmt in ("B_C_D", "BCD"):
            self._write_annexure_b(
                wb, cgrs_sorted, year_new=year_new, year_old=year_old,
                section_name=section_name,
            )
            self._write_annexure_c(
                wb, cgrs_sorted, ffp_by_id, pred_by_id, pipeline=pipeline,
                year_new=year_new, year_old=year_old,
                project=project,
            )
            self._write_annexure_d(
                wb, cgrs_sorted, ffp_by_id, pred_by_id, pipeline=pipeline,
                year_new=year_new, year_old=year_old,
                project=project,
            )
        else:
            raise ValueError(f"unknown format {format!r}; expected 'E_F' or 'B_C_D'")

        if flag_report is not None:
            self._write_issues_sheet(wb, flag_report)

        wb.save(out_path)

    # ------------------------------------------------------------------
    # Annexure E — Run to Run Comparison
    # ------------------------------------------------------------------

    def _write_annexure_e(
        self,
        wb: Workbook,
        cgrs_sorted: list[CGRResult],
        *,
        year_new: str,
        year_old: str,
        section_name: str,
        sheet_name: str = "Annexure E",
        title_text: str | None = None,
    ) -> None:
        # v0.2.5: sheet_name + title_text overrides for the topic-based
        # builder. Defaults preserve legacy E_F dispatch behaviour.
        ws = wb.create_sheet(sheet_name)
        n_cols = len(_ANNEX_E_COLUMNS)
        last_col_letter = get_column_letter(n_cols)

        # Column widths.
        for i, (_label, width, _) in enumerate(_ANNEX_E_COLUMNS, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Row 1 — title (merge A1:Mlast)
        ws.cell(1, 1).value = title_text or "Annexure E: Run to Run Comparison"
        ws.merge_cells(f"A1:{last_col_letter}1")
        _style_cell(ws.cell(1, 1), font=_TITLE_FONT, fill=_TITLE_FILL,
                    alignment=_CENTER, border=_BORDER)
        ws.row_dimensions[1].height = 22

        # Row 2 — group headers
        # Merge A2:D2 ("Feature Detail as per ILI {year_new}")
        # Merge E2:F2 ("Abs. Distance, (m)")
        # Merge G2:H2 ("Anomaly Depth, (%)")
        # Merge I2:J2 ("Anomaly Orientation")
        # Merge K2:L2 ("Anomaly Location")
        # Columns 13 (M) + 14 (N) — no merge; v0.2.6 holds the CGR raw
        # + CGR post-floor pair, each with its own row-3 header.
        ws.cell(2, 1).value = f"Feature Detail as per ILI {year_new}"
        ws.merge_cells("A2:D2")
        ws.cell(2, 5).value = "Abs. Distance, (m)"
        ws.merge_cells("E2:F2")
        ws.cell(2, 7).value = "Anomaly Depth, (%)"
        ws.merge_cells("G2:H2")
        ws.cell(2, 9).value = "Anomaly Orientation"
        ws.merge_cells("I2:J2")
        ws.cell(2, 11).value = "Anomaly Location"
        ws.merge_cells("K2:L2")
        for col in (1, 5, 7, 9, 11):
            _style_cell(ws.cell(2, col), font=_HEADER_FONT, fill=_HEADER_FILL,
                        alignment=_CENTER, border=_BORDER)
        # Also style the unmerged tail cells so the row looks uniform.
        # v0.2.6: cols 13 and 14 are both unmerged CGR columns.
        for col in (2, 3, 4, 6, 8, 10, 12, 13, 14):
            _style_cell(ws.cell(2, col), font=_HEADER_FONT, fill=_HEADER_FILL,
                        alignment=_CENTER, border=_BORDER)
        ws.row_dimensions[2].height = 22

        # Row 3 — field headers (S.N. / Anomaly ID / WT / Joint Number /
        # x4 paired / CGR raw / CGR post-floor).
        headers = [
            "S.N.", "Anomaly ID", "Wall Thickness, (mm)", "Joint Number",
            f"ILI {year_new}", f"ILI {year_old}",
            f"ILI {year_new}", f"ILI {year_old}",
            f"ILI {year_new}", f"ILI {year_old}",
            f"ILI {year_new}", f"ILI {year_old}",
            "CGR raw (mm/yr)",
            "CGR (mm/yr)",
        ]
        for i, label in enumerate(headers, start=1):
            cell = ws.cell(3, i)
            cell.value = label
            _style_cell(cell, font=_HEADER_FONT, fill=_HEADER_FILL,
                        alignment=_CENTER, border=_BORDER)
        ws.row_dimensions[3].height = 22

        # Data rows
        row = 4
        for sn, cgr in enumerate(cgrs_sorted, start=1):
            f_new = cgr.feature
            f_old = cgr.matched_to_run1
            self._fill_annex_e_row(ws, row, sn, cgr, f_new, f_old)
            row += 1

        # Freeze the three header rows.
        ws.freeze_panes = "A4"

    @staticmethod
    def _fill_annex_e_row(
        ws: Worksheet,
        row: int,
        sn: int,
        cgr: CGRResult,
        f_new: Feature,
        f_old: Feature | None,
    ) -> None:
        values = [
            sn,
            f_new.anomaly_id,
            f_new.wt_mm,
            f_new.joint_number,
            f_new.abs_distance_m,
            f_old.abs_distance_m if f_old is not None else None,
            f_new.depth_pct_wt,
            f_old.depth_pct_wt if f_old is not None else cgr.depth_old_used_mm * 100.0 / (f_new.wt_mm or 1.0),
            _format_clock_hhmm(f_new.clock_decimal_hours),
            _format_clock_hhmm(f_old.clock_decimal_hours) if f_old is not None else "",
            _format_surface(f_new.surface),
            _format_surface(f_old.surface) if f_old is not None else "",
            # v0.2.6: pair of CGR columns. `feature_cgr_mm_yr` is the
            # raw pre-floor per-feature rate (matches the BPCL hand
            # formula (d_new − d_unmatched)/Δt); `cgr_mm_yr` is the
            # post-floor value used downstream by FFP projection. In
            # feature_specific / population_only modes the two are
            # identical; in hybrid mode they may differ for features
            # whose raw rate was below the surface P95 floor.
            float(cgr.feature_cgr_mm_yr),
            float(cgr.cgr_mm_yr),
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row, col)
            cell.value = val
            cell.font = _DATA_FONT
            cell.border = _BORDER
            cell.alignment = _CENTER
            cell.number_format = _ANNEX_E_COLUMNS[col - 1][2]

    # ------------------------------------------------------------------
    # Annexure F — Metal Loss Anomalies
    # ------------------------------------------------------------------

    def _write_annexure_f(
        self,
        wb: Workbook,
        cgrs_sorted: list[CGRResult],
        ffp_by_id: dict[str, FFPResult],
        pred_by_id: dict[str, Any],
        *,
        pipeline: Pipeline | None,
        section_name: str,
        pipeline_name: str,
        project: Project | None,
        sheet_name: str = "Annexure F",
        title_text: str | None = None,
    ) -> None:
        # v0.2.5: sheet_name + title_text overrides for the topic-based
        # builder. Defaults preserve legacy E_F dispatch behaviour.
        ws = wb.create_sheet(sheet_name)
        n_cols = len(_ANNEX_F_COLUMNS)
        last_col_letter = get_column_letter(n_cols)

        # Column widths.
        for i, (_label, width, _) in enumerate(_ANNEX_F_COLUMNS, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Row 1 — title
        ws.cell(1, 1).value = title_text or "Annexure F: Metal Loss Anomalies"
        ws.merge_cells(f"A1:{last_col_letter}1")
        _style_cell(ws.cell(1, 1), font=_TITLE_FONT, fill=_TITLE_FILL,
                    alignment=_CENTER, border=_BORDER)
        ws.row_dimensions[1].height = 22

        # Row 2 — pipeline name
        ws.cell(2, 1).value = pipeline_name or ""
        ws.merge_cells(f"A2:{last_col_letter}2")
        _style_cell(ws.cell(2, 1), font=_SUBTITLE_FONT, alignment=_CENTER, border=_BORDER)

        # Row 3 — section name
        ws.cell(3, 1).value = section_name or ""
        ws.merge_cells(f"A3:{last_col_letter}3")
        _style_cell(ws.cell(3, 1), font=_SUBTITLE_FONT, alignment=_CENTER, border=_BORDER)

        # Row 4 — field headers
        for i, (label, _w, _fmt) in enumerate(_ANNEX_F_COLUMNS, start=1):
            cell = ws.cell(4, i)
            cell.value = label
            _style_cell(cell, font=_HEADER_FONT, fill=_HEADER_FILL,
                        alignment=_CENTER, border=_BORDER)
        ws.row_dimensions[4].height = 30

        # Pre-compute horizon end date for "After ..." labels.
        run2_date = (
            project.run_2.inspection_date if (project and project.run_2) else None
        )
        horizon_end = (
            horizon_end_date(run2_date, self.horizon_years) if run2_date else None
        )

        # Data
        row = 5
        for sn, cgr in enumerate(cgrs_sorted, start=1):
            f = cgr.feature
            ffp = ffp_by_id.get(f.anomaly_id)
            pred = pred_by_id.get(f.anomaly_id)

            joint_no = f.joint_number
            joint_length_m = None
            dist_to_weld_m = None
            if pipeline is not None and joint_no is not None:
                j = next((j for j in pipeline.maop_zones), None)  # placeholder
                # Actual joint length: pull from ILIRun.joints isn't directly
                # available here without re-plumbing; leave None and let
                # callers extend if needed.
            joint_length_m = None
            dist_to_weld_m = f.upstream_weld_dist_m

            repair_str = _format_repair_date(pred, horizon_end)

            values = [
                sn,
                f.anomaly_id,
                f.abs_distance_m,
                f.latitude,
                f.longitude,
                joint_no,
                joint_length_m,
                dist_to_weld_m,
                f.raw_description or "",
                _format_surface(f.surface),
                f.wt_mm,
                _format_clock_hhmm(f.clock_decimal_hours),
                f.depth_pct_wt,
                f.length_mm,
                f.width_mm,
                repair_str,
            ]
            for col, val in enumerate(values, start=1):
                cell = ws.cell(row, col)
                cell.value = val
                cell.font = _DATA_FONT
                cell.border = _BORDER
                cell.alignment = _CENTER
                cell.number_format = _ANNEX_F_COLUMNS[col - 1][2]
            row += 1

        ws.freeze_panes = "A5"

    # ------------------------------------------------------------------
    # Annexure B — older format, run-to-run with CGR
    # ------------------------------------------------------------------

    def _write_annexure_b(
        self,
        wb: Workbook,
        cgrs_sorted: list[CGRResult],
        *,
        year_new: str,
        year_old: str,
        section_name: str,
    ) -> None:
        ws = wb.create_sheet("Annexure B")
        cols = ["S.N.", "Feature ID", "Joint number", "Absolute distance",
                "Wall Thickness", "Depth new", "Depth old",
                "Surface new", "Surface old", "CGR"]
        # Widths
        for i, w in enumerate([8, 12, 12, 16, 12, 12, 12, 12, 12, 12], start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        last_col_letter = get_column_letter(len(cols))

        ws.cell(1, 2).value = "ANNEXURE -  B"
        ws.merge_cells("B1:I1")
        _style_cell(ws.cell(1, 2), font=_TITLE_FONT, fill=_TITLE_FILL,
                    alignment=_CENTER, border=_BORDER)

        ws.cell(2, 1).value = f"Results of ILI Comparison ({section_name}) — matched defects"
        ws.merge_cells(f"A2:{last_col_letter}3")
        _style_cell(ws.cell(2, 1), font=_SUBTITLE_FONT, alignment=_CENTER, border=_BORDER)

        # Row 6 — group headers
        ws.cell(6, 1).value = "S.N."
        ws.cell(6, 2).value = "Feature ID"
        ws.cell(6, 3).value = "Joint number"
        ws.cell(6, 4).value = "Absolute distance"
        ws.cell(6, 5).value = "Wall Thickness"
        ws.cell(6, 6).value = "Feature Depth, %"
        ws.merge_cells("F6:G6")
        ws.cell(6, 8).value = "Wall side"
        ws.merge_cells("H6:I6")
        ws.cell(6, 10).value = "CGR"
        for c in range(1, len(cols) + 1):
            _style_cell(ws.cell(6, c), font=_HEADER_FONT, fill=_HEADER_FILL,
                        alignment=_CENTER, border=_BORDER)

        # Row 7 — sub-header
        sub = ["", "", "", "mtrs.", "mm",
               f"ILI {year_new}", f"ILI {year_old}",
               f"ILI {year_new}", f"ILI {year_old}", "mm/yr"]
        for i, v in enumerate(sub, start=1):
            cell = ws.cell(7, i)
            cell.value = v
            _style_cell(cell, font=_HEADER_FONT, fill=_HEADER_FILL,
                        alignment=_CENTER, border=_BORDER)

        # Data — only matched features (Annexure B is the "matched" sheet)
        row = 8
        sn = 0
        for cgr in cgrs_sorted:
            if cgr.matched_to_run1 is None:
                continue
            sn += 1
            f_new = cgr.feature
            f_old = cgr.matched_to_run1
            values = [
                sn,
                f_new.anomaly_id,
                f_new.joint_number,
                f_new.abs_distance_m,
                f_new.wt_mm,
                f_new.depth_pct_wt,
                f_old.depth_pct_wt,
                _format_surface(f_new.surface),
                _format_surface(f_old.surface),
                float(cgr.cgr_mm_yr),
            ]
            formats = ["0", "@", "0", "0.000", "0.0", "0.00", "0.00", "@", "@", "0.0000"]
            for col, (val, fmt) in enumerate(zip(values, formats), start=1):
                cell = ws.cell(row, col)
                cell.value = val
                cell.font = _DATA_FONT
                cell.border = _BORDER
                cell.alignment = _CENTER
                cell.number_format = fmt
            row += 1
        ws.freeze_panes = "A8"

    # ------------------------------------------------------------------
    # Annexure C — B31G for ALL defects (current + 10-year)
    # ------------------------------------------------------------------

    def _write_annexure_c(
        self,
        wb: Workbook,
        cgrs_sorted: list[CGRResult],
        ffp_by_id: dict[str, FFPResult],
        pred_by_id: dict[str, Any],
        *,
        pipeline: Pipeline | None,
        year_new: str,
        year_old: str,
        project: Project | None,
    ) -> None:
        self._write_bcd_assessment_sheet(
            wb, "Annexure C", "ASME B-31G Original",
            cgrs_sorted, ffp_by_id, pred_by_id,
            pipeline=pipeline, project=project,
            method_for_psafe="ASME B-31G Original",
        )

    # ------------------------------------------------------------------
    # Annexure D — Kastner for circumferential defects
    # ------------------------------------------------------------------

    def _write_annexure_d(
        self,
        wb: Workbook,
        cgrs_sorted: list[CGRResult],
        ffp_by_id: dict[str, FFPResult],
        pred_by_id: dict[str, Any],
        *,
        pipeline: Pipeline | None,
        year_new: str,
        year_old: str,
        project: Project | None,
    ) -> None:
        # Filter to only features where Kastner was the controlling method
        # (circumferential defects).
        circ_cgrs = [
            c for c in cgrs_sorted
            if (
                ffp_by_id.get(c.feature.anomaly_id) is not None
                and ffp_by_id[c.feature.anomaly_id].method is FFPMethod.KASTNER
            )
        ]
        self._write_bcd_assessment_sheet(
            wb, "Annexure D", "Kastner Approach",
            circ_cgrs, ffp_by_id, pred_by_id,
            pipeline=pipeline, project=project,
            method_for_psafe="Kastner Approach",
        )

    @staticmethod
    def _write_bcd_assessment_sheet(
        wb: Workbook,
        sheet_name: str,
        method_label: str,
        cgrs_sorted: list[CGRResult],
        ffp_by_id: dict[str, FFPResult],
        pred_by_id: dict[str, Any],
        *,
        pipeline: Pipeline | None,
        project: Project | None,
        method_for_psafe: str,
        title_text: str | None = None,
    ) -> None:
        """Shared 4-row-header assessment sheet used by Annexures C and D
        in the older format. Columns:
            S.N. | Feature ID | Joint No. | Chainage | Surface |
            Depth (now / +10) | SOP (now / +10) | ERF (now / +10) |
            CGR (mm/yr) | Repair Date
        """
        ws = wb.create_sheet(sheet_name)
        widths = [8, 12, 12, 14, 10, 12, 12, 14, 14, 12, 12, 12, 14]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        last_col = len(widths)
        last_col_letter = get_column_letter(last_col)

        # Row 1 — title (v0.2.5: title_text overrides the default ANNEXURE-X
        # banner; defaults to the legacy "ANNEXURE -  X" wording when called
        # via the legacy B_C_D dispatch).
        ws.cell(1, 1).value = (
            title_text or f"ANNEXURE -  {sheet_name.split()[-1]}"
        )
        ws.merge_cells(f"A1:{last_col_letter}1")
        _style_cell(ws.cell(1, 1), font=_TITLE_FONT, fill=_TITLE_FILL,
                    alignment=_CENTER, border=_BORDER)

        # Row 2 — subtitle
        ws.cell(2, 1).value = f"Estimated ERF of Defects Reported in ILI — {method_label}"
        ws.merge_cells(f"A2:{last_col_letter}2")
        _style_cell(ws.cell(2, 1), font=_SUBTITLE_FONT, alignment=_CENTER, border=_BORDER)

        # Row 3 — group headers
        ws.cell(3, 6).value = "Feature Depth (%)"
        ws.merge_cells("F3:G3")
        ws.cell(3, 8).value = f"SOP (Kg/cm2) - {method_label}"
        ws.merge_cells("H3:I3")
        ws.cell(3, 10).value = f"ERF {method_label}"
        ws.merge_cells("J3:K3")
        for col in (6, 8, 10):
            _style_cell(ws.cell(3, col), font=_HEADER_FONT, fill=_HEADER_FILL,
                        alignment=_CENTER, border=_BORDER)

        # Row 4 — field headers
        run2_date = project.run_2.inspection_date if (project and project.run_2) else None
        horizon_end = (
            horizon_end_date(run2_date, 10) if run2_date else None
        )
        date_now = run2_date.isoformat() if run2_date else "Now"
        date_h10 = horizon_end.isoformat() if horizon_end else "+10 yr"

        h4 = ["S.N.", "Feature ID", "Joint No.", "Chainage-(mtrs)", "Surface",
              date_now, date_h10,
              date_now, date_h10,
              date_now, date_h10,
              "CGR- mm/yr", "Repair Date"]
        for i, label in enumerate(h4, start=1):
            cell = ws.cell(4, i)
            cell.value = label
            _style_cell(cell, font=_HEADER_FONT, fill=_HEADER_FILL,
                        alignment=_CENTER, border=_BORDER)
        ws.row_dimensions[4].height = 22

        # Data
        row = 5
        for sn, cgr in enumerate(cgrs_sorted, start=1):
            f = cgr.feature
            ffp = ffp_by_id.get(f.anomaly_id)
            pred = pred_by_id.get(f.anomaly_id)
            # Year-now / +10 values from the prediction (when available).
            depth_now = f.depth_pct_wt or 0.0
            psafe_now = ffp.sop_kgcm2 if ffp else 0.0
            erf_now = ffp.erf if ffp else 0.0

            if pred is not None and pred.yearly_assessments:
                # The last yearly assessment is the horizon-end state.
                last = pred.yearly_assessments[-1]
                depth_h10 = last.depth_pct_wt
                psafe_h10 = last.sop_kgcm2
                erf_h10 = last.erf
            else:
                depth_h10 = depth_now
                psafe_h10 = psafe_now
                erf_h10 = erf_now

            repair_str = _format_repair_date(pred, horizon_end)

            values = [
                sn,
                f.anomaly_id,
                f.joint_number,
                f.abs_distance_m,
                _format_surface(f.surface),
                depth_now,
                depth_h10,
                psafe_now,
                psafe_h10,
                erf_now,
                erf_h10,
                float(cgr.cgr_mm_yr),
                repair_str,
            ]
            formats = ["0", "@", "0", "0.000", "@",
                       "0.00", "0.00",
                       "0.00", "0.00",
                       "0.0000", "0.0000",
                       "0.0000", "@"]
            for col, (val, fmt) in enumerate(zip(values, formats), start=1):
                cell = ws.cell(row, col)
                cell.value = val
                cell.font = _DATA_FONT
                cell.border = _BORDER
                cell.alignment = _CENTER
                cell.number_format = fmt
            row += 1
        ws.freeze_panes = "A5"

    # ------------------------------------------------------------------
    # QA "Issues" sheet — appended when a FlagReport is supplied
    # ------------------------------------------------------------------

    @staticmethod
    def _write_issues_sheet(
        wb: Workbook,
        flag_report: Any,
        *,
        sheet_name: str = "QA Issues",
        title_text: str | None = None,
    ) -> None:
        # v0.2.5: sheet_name + title_text overrides for topic-based builder.
        ws = wb.create_sheet(sheet_name)
        widths = [10, 32, 14, 12, 60]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.cell(1, 1).value = title_text or "QA Issues"
        ws.merge_cells("A1:E1")
        _style_cell(ws.cell(1, 1), font=_TITLE_FONT, fill=_TITLE_FILL,
                    alignment=_CENTER, border=_BORDER)

        ws.cell(2, 1).value = getattr(flag_report, "summary", "") or ""
        ws.merge_cells("A2:E2")
        _style_cell(ws.cell(2, 1), font=_SUBTITLE_FONT, alignment=_LEFT, border=_BORDER)

        for i, label in enumerate(
            ["Severity", "Code", "Feature", "Source row", "Message"], start=1
        ):
            cell = ws.cell(3, i)
            cell.value = label
            _style_cell(cell, font=_HEADER_FONT, fill=_HEADER_FILL,
                        alignment=_CENTER, border=_BORDER)

        row = 4
        flags = getattr(flag_report, "all_flags", []) or []
        # Stable order: ERROR -> WARN -> INFO, then by code.
        sev_order = {"error": 0, "warn": 1, "info": 2}
        flags_sorted = sorted(
            flags,
            key=lambda f: (
                sev_order.get(getattr(f.severity, "value", ""), 9),
                getattr(f.code, "value", ""),
            ),
        )
        for f in flags_sorted:
            ws.cell(row, 1).value = getattr(f.severity, "value", "").upper()
            ws.cell(row, 2).value = getattr(f.code, "value", "")
            ws.cell(row, 3).value = f.feature_id or ""
            ws.cell(row, 4).value = f.source_row if f.source_row is not None else ""
            ws.cell(row, 5).value = f.message
            for c in range(1, 6):
                cell = ws.cell(row, c)
                cell.font = _DATA_FONT
                cell.border = _BORDER
                cell.alignment = _LEFT
            row += 1


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _style_cell(
    cell,
    *,
    font: Font | None = None,
    fill: PatternFill | None = None,
    alignment: Alignment | None = None,
    border: Border | None = None,
) -> None:
    if font is not None:
        cell.font = font
    if fill is not None:
        cell.fill = fill
    if alignment is not None:
        cell.alignment = alignment
    if border is not None:
        cell.border = border


def _format_clock_hhmm(clock_decimal_hours: float | None) -> str:
    """Render decimal hours as 'hh:mm:ss' to match the published format."""
    if clock_decimal_hours is None:
        return ""
    hh = int(clock_decimal_hours) % 12
    rem = (clock_decimal_hours - int(clock_decimal_hours)) * 60.0
    mm = int(rem)
    ss = int(round((rem - mm) * 60.0))
    if ss == 60:
        ss = 0
        mm += 1
    if mm == 60:
        mm = 0
        hh = (hh + 1) % 12
    return f"{hh:02d}:{mm:02d}:{ss:02d}"


def _format_surface(surface: Surface | None) -> str:
    """Render surface as 'int.' / 'ext.' / '' to match the published format."""
    if surface is None:
        return ""
    if surface is Surface.INTERNAL:
        return "int."
    if surface is Surface.EXTERNAL:
        return "ext."
    if surface is Surface.MIDWALL:
        return "mid."
    return ""


def _format_repair_date(pred: Any, horizon_end: date | None) -> str:
    """Render the predicted-repair-year cell.

    Matched to the conventions in the reference files:
      * If the prediction triggered within the horizon → 'dd-mm-yyyy'
        of the predicted repair date.
      * If not → 'After {Month Year}' string using the horizon end.
    """
    if pred is None:
        return ""
    trigger = getattr(pred, "repair_trigger", TRIGGER_NONE)
    repair_date = getattr(pred, "predicted_repair_date", None)
    if trigger in (TRIGGER_DEPTH_80, TRIGGER_ERF_1) and repair_date is not None:
        return repair_date.strftime("%d-%m-%Y")
    # NONE_WITHIN_HORIZON or no run-2 date supplied
    if horizon_end is not None:
        return f"After {horizon_end.strftime('%B %Y')}"
    return "After horizon"


def _resolve_year_labels(project: Project | None) -> tuple[str, str]:
    """Pull 4-digit year labels from project.run_1 / run_2 inspection_date.

    Falls back to 'run 2' / 'run 1' strings if dates aren't set.
    """
    if project is None:
        return ("run 2", "run 1")
    run_2_date = getattr(getattr(project, "run_2", None), "inspection_date", None)
    run_1_date = getattr(getattr(project, "run_1", None), "inspection_date", None)
    new = str(run_2_date.year) if run_2_date else "run 2"
    old = str(run_1_date.year) if run_1_date else "run 1"
    return (new, old)


def _resolve_section_labels(
    project: Project | None,
    pipeline: Pipeline | None,
) -> tuple[str, str]:
    section = ""
    pname = ""
    if project is not None:
        pname = project.project_name or ""
        if pipeline is None:
            pipeline = project.pipeline
    if pipeline is not None:
        section = pipeline.pipeline_name or section
        pname = pname or pipeline.client_name or pipeline.pipeline_name or ""
    return (section, pname)


# ---------------------------------------------------------------------------
# v0.2.5 — new topic writers + sheet-name helper
# ---------------------------------------------------------------------------

_SECTION_HEADER_FILL = PatternFill(
    start_color="D9D9D9", end_color="D9D9D9", fill_type="solid",
)
_SECTION_HEADER_FONT = Font(bold=True, size=11)
_PLACEHOLDER_FILL = PatternFill(
    start_color="FFF2CC", end_color="FFF2CC", fill_type="solid",
)

# Excel sheet-name hard cap.
_MAX_SHEET_NAME_LEN = 31


def make_topic_sheet_name(letter: str, display_name: str) -> tuple[str, str]:
    """Compute the sheet-tab name + full row-1 title for a topic.

    Target form: ``"Annexure {letter} — {display_name}"``. If that
    exceeds Excel's 31-char limit, fall back to ``"Annexure {letter}"``
    as the sheet name while keeping the full string as the row-1 title.

    Returns ``(sheet_name, title_text)``.
    """
    full = f"Annexure {letter} — {display_name}"
    if len(full) <= _MAX_SHEET_NAME_LEN:
        return (full, full)
    return (f"Annexure {letter}"[:_MAX_SHEET_NAME_LEN], full)


def _kv_row(
    ws: Worksheet,
    row: int,
    key: str,
    value: Any,
    *,
    key_col: int = 1,
    value_col: int = 2,
) -> None:
    """Render a two-column "key | value" row in the Guidelines sheet."""
    k = ws.cell(row, key_col)
    k.value = key
    k.font = _DATA_FONT
    k.border = _BORDER
    k.alignment = _LEFT
    v = ws.cell(row, value_col)
    v.value = "" if value is None else value
    v.font = _DATA_FONT
    v.border = _BORDER
    v.alignment = _LEFT


def _section_header(
    ws: Worksheet,
    row: int,
    text: str,
    *,
    n_cols: int = 4,
) -> None:
    """Render a section header — bold, light-gray fill, merged across cols."""
    ws.cell(row, 1).value = text
    last = get_column_letter(n_cols)
    ws.merge_cells(f"A{row}:{last}{row}")
    _style_cell(
        ws.cell(row, 1),
        font=_SECTION_HEADER_FONT,
        fill=_SECTION_HEADER_FILL,
        alignment=_LEFT,
        border=_BORDER,
    )
    ws.row_dimensions[row].height = 20


# v0.2.5 — formula text per FFP method (hardcoded; do NOT parse
# docs/ENGINE_REFERENCE.md at runtime).
_FFP_FORMULA_TEXT: dict[str, str] = {
    "B31G_Original": (
        "ASME B31G-2012 Section 4 (Level 1/2). "
        "z = L²/(D·t); S_flow = 1.1·SMYS; "
        "Low-z (z ≤ 20): M = √(1 + 0.8·z); Q = (2/3)·(d/t); "
        "R = (1 − Q)/(1 − Q/M). "
        "High-z (z > 20): R = 1 − d/t. "
        "Pf = 2·S_flow·t/D · R; Psafe = Pf · Fd; ERF = MAOP/Psafe."
    ),
    "B31G_Modified": (
        "ASME B31G-2012 Section 5 (0.85·dL method). "
        "z = L²/(D·t); S_flow = SMYS + 69 MPa (≈ SMYS + 10 ksi); "
        "Low-z (z ≤ 50): M = √(1 + 0.6275·z − 0.003375·z²). "
        "High-z (z > 50): M = 0.032·z + 3.3. "
        "Q = 0.85·d/t; SF = S_flow·(1 − Q)/(1 − Q/M); "
        "Pf = 2·SF·t/D; Psafe = Pf · Fd; ERF = MAOP/Psafe."
    ),
    "RSTRENG": (
        "PRCI RSTRENG. Without river-bottom depth profile (typical "
        "for POF 110 vendor data) this falls back to the B31G "
        "Modified 0.85·dL form — Pf and Psafe are identical to "
        "B31G_Modified. The `using_approximate_profile=True` flag on "
        "each FFPResult records the limitation."
    ),
    "DNV_RP_F101": (
        "DNV-RP-F101 (2017) Part B, ASD format. "
        "z = L²/(D·t); Q = √(1 + 0.31·z); "
        "Pf = 2·UTS·t/(D−t) · (1 − d/t)/(1 − (d/t)/Q); "
        "Psafe = Pf · Fd. UTS, not flow stress. If UTS not supplied, "
        "estimated as SMYS + 110 MPa (standard line-pipe correlation)."
    ),
    "Kastner": (
        "Net-section approximation for partial-depth circumferential "
        "defects (deferred-implementation simplification of Kastner "
        "1986). S_flow = 1.1·SMYS; α = (W/(π·D))·(d/t), clamped [0,1]; "
        "Pf = 4·S_flow·t/D · (1 − α); Psafe = Pf · Fd."
    ),
}


def _write_guidelines_formulas_sheet(
    wb: Workbook,
    sheet_name: str,
    project: Project,
    results: Any,
    run2_year: int | None,
    *,
    title_text: str | None = None,
) -> None:
    """Topic 1: Guidelines & Formulas Used.

    Produces a single sheet with 9 sections — project header, pipeline
    parameters, MAOP zones, ILI runs, FFP method, CGR computation,
    repair prediction, critical constants, and references.

    Two-column key/value layout (with a 4-col table for MAOP zones).
    """
    from datetime import date as _date

    ws = wb.create_sheet(sheet_name)
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 16

    # Row 1: title (merged A1:D1, yellow fill, bold 12pt — matches
    # Annexure E header style).
    ws.cell(1, 1).value = (
        title_text or "Annexure A — Guidelines & Formulas Used"
    )
    ws.merge_cells("A1:D1")
    _style_cell(
        ws.cell(1, 1),
        font=_TITLE_FONT,
        fill=_TITLE_FILL,
        alignment=_CENTER,
        border=_BORDER,
    )
    ws.row_dimensions[1].height = 22

    row = 3

    # ----- Section 1: Project Header --------------------------------------
    _section_header(ws, row, "Section 1 — Project Header")
    row += 1
    _kv_row(ws, row, "project_name", project.project_name or ""); row += 1
    _kv_row(ws, row, "pipeline_name",
            project.pipeline.pipeline_name if project.pipeline else ""); row += 1
    _kv_row(ws, row, "client_name",
            project.pipeline.client_name if project.pipeline else ""); row += 1
    _kv_row(ws, row, "report_number", project.report_number or ""); row += 1
    _kv_row(ws, row, "report_revision", project.report_revision or ""); row += 1
    _kv_row(ws, row, "prepared_by", project.prepared_by or ""); row += 1
    _kv_row(ws, row, "generation_date", _date.today().isoformat()); row += 1
    row += 1

    # ----- Section 2: Pipeline Parameters ---------------------------------
    p = project.pipeline
    _section_header(ws, row, "Section 2 — Pipeline Parameters"); row += 1
    _kv_row(ws, row, "diameter_mm",   p.diameter_mm   if p else 0); row += 1
    _kv_row(ws, row, "length_km",     p.length_km     if p else 0); row += 1
    _kv_row(ws, row, "material_grade", p.material_grade if p else ""); row += 1
    _kv_row(ws, row, "smys_mpa",      p.smys_mpa      if p else 0); row += 1
    _kv_row(ws, row, "install_year",  p.install_year  if p else 0); row += 1
    _kv_row(ws, row, "product",       p.product       if p else ""); row += 1
    _kv_row(ws, row, "service_class", p.service_class if p else ""); row += 1
    row += 1

    # ----- Section 3: MAOP Zones (4-col table) ----------------------------
    # v0.3.0: section header includes the zoning mode; the bound column
    # switches between "WT range (mm)" and "Chainage range (m)" based
    # on the pipeline's `maop_zoning_mode`.
    mode = getattr(p, "maop_zoning_mode", "wt") if p else "wt"
    section_label = "Section 3 — MAOP Zones"
    if mode == "chainage":
        section_label += " (chainage-bounded)"
    _section_header(ws, row, section_label); row += 1
    # Header row — bound column changes with mode.
    bound_header = (
        "Chainage range (m)" if mode == "chainage" else "WT range (mm)"
    )
    headers = ["Zone", bound_header, "Design factor", "MAOP (kg/cm²)"]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row, i)
        cell.value = h
        _style_cell(
            cell, font=_HEADER_FONT, fill=_HEADER_FILL,
            alignment=_CENTER, border=_BORDER,
        )
    row += 1
    zones = (p.maop_zones if p else []) or []
    for i, z in enumerate(zones, start=1):
        ws.cell(row, 1).value = f"Z{i}"
        if mode == "chainage":
            lo = z.chainage_m_min if z.chainage_m_min is not None else 0.0
            hi = z.chainage_m_max if z.chainage_m_max is not None else 0.0
            ws.cell(row, 2).value = f"{lo:.1f} – {hi:.1f}"
        else:
            lo = z.wt_mm_min if z.wt_mm_min is not None else 0.0
            hi = z.wt_mm_max if z.wt_mm_max is not None else 0.0
            ws.cell(row, 2).value = f"{lo:.1f} – {hi:.1f}"
        ws.cell(row, 3).value = z.design_factor
        ws.cell(row, 4).value = z.maop_kgcm2
        for c in range(1, 5):
            cell = ws.cell(row, c)
            cell.font = _DATA_FONT
            cell.border = _BORDER
            cell.alignment = _CENTER
        ws.cell(row, 3).number_format = "0.00"
        ws.cell(row, 4).number_format = "0.0"
        row += 1
    if not zones:
        ws.cell(row, 1).value = "(no zones declared)"
        row += 1
    row += 1

    # ----- Section 4: ILI Runs --------------------------------------------
    _section_header(ws, row, "Section 4 — ILI Runs"); row += 1
    r1 = project.run_1
    r2 = project.run_2
    _kv_row(ws, row, "Run-1 inspection date",
            r1.inspection_date.isoformat() if r1.inspection_date else ""); row += 1
    _kv_row(ws, row, "Run-1 vendor", r1.vendor or ""); row += 1
    _kv_row(ws, row, "Run-1 tool_type", r1.tool_type or ""); row += 1
    _kv_row(ws, row, "Run-1 file_path", r1.file_path or ""); row += 1
    _kv_row(ws, row, "Run-2 inspection date",
            r2.inspection_date.isoformat() if r2.inspection_date else ""); row += 1
    _kv_row(ws, row, "Run-2 vendor", r2.vendor or ""); row += 1
    _kv_row(ws, row, "Run-2 tool_type", r2.tool_type or ""); row += 1
    _kv_row(ws, row, "Run-2 file_path", r2.file_path or ""); row += 1
    years_between = getattr(results, "years_between", 0.0) or project.years_between_runs
    _kv_row(ws, row, "Δt (years between runs)",
            f"{years_between:.6f}"); row += 1
    row += 1

    # ----- Section 5: FFP Method ------------------------------------------
    cfg = project.config or {}
    ffp_cfg = (cfg.get("ffp") or {})
    primary = ffp_cfg.get("primary_method", "B31G_Original")
    _section_header(ws, row, "Section 5 — FFP Method"); row += 1
    _kv_row(ws, row, "primary_method", primary); row += 1
    _kv_row(ws, row, "Formula",
            _FFP_FORMULA_TEXT.get(primary, "(formula not registered)")); row += 1
    _kv_row(ws, row, "cross_check_methods",
            ", ".join(ffp_cfg.get("cross_check_methods") or []) or "(none)"); row += 1
    _kv_row(ws, row, "kastner_for_circumferential",
            ffp_cfg.get("kastner_for_circumferential", True)); row += 1
    row += 1

    # ----- Section 6: CGR Computation -------------------------------------
    cgr_cfg = (cfg.get("cgr") or {})
    _section_header(ws, row, "Section 6 — CGR Computation"); row += 1
    _kv_row(ws, row, "mode", cgr_cfg.get("mode", "hybrid")); row += 1
    _kv_row(ws, row, "split_by_surface",
            cgr_cfg.get("split_by_surface", True)); row += 1
    _kv_row(ws, row, "floor_negative_at_zero",
            cgr_cfg.get("floor_negative_at_zero", True)); row += 1
    _kv_row(ws, row, "unmatched_depth_assumption_pct_wt",
            cgr_cfg.get("unmatched_depth_assumption_pct_wt", 10.0)); row += 1
    _kv_row(ws, row, "extreme_cgr_threshold_mm_yr",
            cgr_cfg.get("extreme_cgr_threshold_mm_yr", 1.0)); row += 1
    _kv_row(ws, row, "tool_depth_tolerance_pct_wt",
            cgr_cfg.get("tool_depth_tolerance_pct_wt", 10.0)); row += 1
    row += 1

    # ----- Section 7: Repair Prediction -----------------------------------
    rp_cfg = (cfg.get("repair_prediction") or {})
    _section_header(ws, row, "Section 7 — Repair Prediction"); row += 1
    _kv_row(ws, row, "horizon_years",
            int(rp_cfg.get("horizon_years", 10))); row += 1
    _kv_row(ws, row, "depth_trigger_pct_wt",
            rp_cfg.get("depth_trigger_pct_wt", 80.0)); row += 1
    _kv_row(ws, row, "erf_trigger",
            rp_cfg.get("erf_trigger", 1.0)); row += 1
    _kv_row(ws, row, "Repair date arithmetic",
            "run2_date + int(N × 365.25) days  (integer year resolution)"); row += 1
    row += 1

    # ----- Section 8: Critical Constants ----------------------------------
    _section_header(ws, row, "Section 8 — Critical Constants"); row += 1
    _kv_row(ws, row, "MPA_TO_KGCM2",
            "10.197162129779283  (exact, = 1 000 000 / 98 066.5)"); row += 1
    _kv_row(ws, row, "S_flow (B31G Original / RSTRENG fallback)",
            "1.1 × SMYS  (MPa)"); row += 1
    _kv_row(ws, row, "S_flow (B31G Modified)",
            "SMYS + 69 MPa  (≈ SMYS + 10 ksi)"); row += 1
    _kv_row(ws, row, "S_flow (DNV-RP-F101)",
            "UTS  (estimated SMYS + 110 MPa if not supplied)"); row += 1
    _kv_row(ws, row, "Folias M branch split (B31G Original)",
            "z = 20  (low_z parabolic / high_z rectangular)"); row += 1
    _kv_row(ws, row, "Folias M branch split (B31G Modified)",
            "z = 50  (polynomial / linear extrapolation)"); row += 1
    row += 1

    # ----- Section 9: References ------------------------------------------
    _section_header(ws, row, "Section 9 — References"); row += 1
    _kv_row(ws, row, "ASME B31G-2012",
            "Manual for Determining the Remaining Strength of Corroded Pipelines"); row += 1
    _kv_row(ws, row, "ASME B31.8",
            "Gas Transmission and Distribution Piping Systems"); row += 1
    _kv_row(ws, row, "DNV-RP-F101 (2017)",
            "Corroded Pipelines"); row += 1
    _kv_row(ws, row, "POF-100 / POF-110",
            "Pipeline Operators Forum — Specifications and Requirements for ILI"); row += 1
    _kv_row(ws, row, "Kiefner & Vieth (1989)",
            "A Modified Criterion for Evaluating the Remaining Strength of "
            "Corroded Pipe (RSTRENG)"); row += 1


def _write_dent_strain_sheet(
    wb: Workbook,
    sheet_name: str,
    project: Project,
    results: Any,
    run2_year: int | None,
    *,
    title_text: str | None = None,
) -> None:
    """Topic 6: Dent strain analysis per ASME B31.8 §851.4.1 / Appendix R.

    v0.3.1: full strain implementation (replaces the v0.2.5 placeholder).
    Reads dent features via the auxiliary path in
    :mod:`src.io.feature_reader` (which bypasses the FFP-pipeline
    skip-list — the metal-loss read path continues to filter dents).
    Per dent, computes ε_1 (circumferential bending), ε_2 (longitudinal
    bending), ε_3 (extensional/membrane), ε_i (inside-surface effective
    strain), ε_o (outside-surface effective strain), and the resultant
    max(|ε_i|, |ε_o|) × 100 in percent.

    Layout matches BPCL's published Annexure E reference (81 dents):
    Feature ID, chainage, joint, geometry, OD, WT, pipe radius, all
    five strain components, resultant, flags.

    Empty-state ("No dent features identified in Run-2.") preserved
    from v0.2.5 for files without dent rows.
    """
    from src.core.dent_strain import compute_dent_strain_from_feature
    from src.io.feature_reader import read_dent_features

    ws = wb.create_sheet(sheet_name)
    columns = [
        ("S.N.",                 6,  "0"),
        ("Feature ID",           12, "@"),
        ("Chainage (m)",         14, "0.000"),
        ("Joint No.",            10, "0"),
        ("Surface",              10, "@"),
        ("Orientation",          12, "@"),
        ("L (mm)",               10, "0"),
        ("W (mm)",               10, "0"),
        ("Depth (mm)",           12, "0.000"),
        ("OD (mm)",              10, "0.0"),
        ("WT (mm)",              10, "0.0"),
        ("Pipe Radius (mm)",     14, "0.00"),
        ("E1",                   12, "0.000000"),
        ("E2",                   12, "0.000000"),
        ("E3",                   12, "0.000000"),
        ("Ei",                   12, "0.000000"),
        ("Eo",                   12, "0.000000"),
        ("Resultant Strain %",   14, "0.0000"),
        ("Flags",                26, "@"),
    ]
    for i, (_label, width, _fmt) in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    n_cols = len(columns)
    last_col_letter = get_column_letter(n_cols)

    # Row 1: title.
    ws.cell(1, 1).value = (
        title_text or "Annexure F — Estimated Strain in Dents per ASME B31.8"
    )
    ws.merge_cells(f"A1:{last_col_letter}1")
    _style_cell(
        ws.cell(1, 1), font=_TITLE_FONT, fill=_TITLE_FILL,
        alignment=_CENTER, border=_BORDER,
    )
    ws.row_dimensions[1].height = 22

    # Row 2: subtitle citing the standard reference.
    ws.cell(2, 1).value = (
        "ASME B31.8 Appendix R / §851.4.1 — ε_1 (circumferential "
        "bending), ε_2 (longitudinal bending), ε_3 (longitudinal "
        "membrane), ε_i / ε_o effective strain at inside / outside "
        "surfaces (Lukasiewicz-Czyz combined-strain form). "
        "Resultant strain ≥ 6 % triggers HIGH_STRAIN_REJECT_CRITERIA."
    )
    ws.merge_cells(f"A2:{last_col_letter}2")
    _style_cell(
        ws.cell(2, 1), font=_SUBTITLE_FONT,
        alignment=Alignment(horizontal="left", vertical="center",
                            wrap_text=True),
        border=_BORDER,
    )
    ws.row_dimensions[2].height = 32

    # Row 3: column headers.
    header_row = 3
    for i, (label, _w, _fmt) in enumerate(columns, start=1):
        cell = ws.cell(header_row, i)
        cell.value = label
        _style_cell(
            cell, font=_HEADER_FONT, fill=_HEADER_FILL,
            alignment=_CENTER, border=_BORDER,
        )
    ws.row_dimensions[header_row].height = 22

    # Dent inventory: v0.3.1 uses src/io/feature_reader.py which
    # bypasses the FFP read-path skip-list specifically for dents.
    run2_path = project.run_2.file_path if project.run_2 else None
    dents: list = read_dent_features(run2_path) if run2_path else []

    # Row 4+: data.
    data_row = header_row + 1
    if not dents:
        ws.cell(data_row, 1).value = "No dent features identified in Run-2."
        ws.merge_cells(f"A{data_row}:{last_col_letter}{data_row}")
        _style_cell(
            ws.cell(data_row, 1), font=_DATA_FONT,
            alignment=Alignment(horizontal="center", vertical="center"),
            border=_BORDER,
        )
        return

    # Sort dents by chainage for stable output (mirrors Annexure E sort).
    dents.sort(key=lambda f: (f.abs_distance_m or 0.0, str(f.anomaly_id)))

    pipeline = getattr(project, "pipeline", None)
    od_mm = getattr(pipeline, "diameter_mm", 0.0) if pipeline else 0.0

    for sn, f in enumerate(dents, start=1):
        r = compute_dent_strain_from_feature(f, pipeline)
        flags_text = " · ".join(r.flags) if r.flags else ""
        values = [
            sn, r.feature_id, r.chainage_m, r.joint_no,
            _format_surface(f.surface),
            _format_clock_hhmm(f.clock_decimal_hours),
            r.length_mm, r.width_mm, r.depth_mm,
            od_mm, r.wt_mm, r.pipe_radius_mm,
            r.E1, r.E2, r.E3, r.Ei, r.Eo,
            r.resultant_strain_pct,
            flags_text,
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(data_row, col)
            cell.value = val
            cell.font = _DATA_FONT
            cell.border = _BORDER
            cell.alignment = _CENTER
            cell.number_format = columns[col - 1][2]
        # Highlight the row if it crosses the 6% reject threshold.
        if "HIGH_STRAIN_REJECT_CRITERIA" in r.flags:
            for col in range(1, n_cols + 1):
                ws.cell(data_row, col).fill = _PLACEHOLDER_FILL
        data_row += 1


# Backward-compat alias: v0.2.5 / v0.3.0 callers that imported
# `_write_dent_strain_placeholder_sheet` continue to work. The
# function is now the real-math implementation regardless.
_write_dent_strain_placeholder_sheet = _write_dent_strain_sheet
