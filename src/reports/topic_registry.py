"""Annexure topic registry (v0.2.5).

v0.2.0–v0.2.4 picked the report layout via a single ``annexure_format``
string — ``"E_F"`` or ``"B_C_D"`` — locked to two presets. Each preset
spawned a fixed set of sheets in a fixed order with fixed annexure
letters. That doesn't fit customers who want, say, "the Guidelines
sheet from preset A plus the ERF table from preset B, but the
Kastner sheet labelled D2 instead of D".

v0.2.5 replaces the preset selector with a **topic registry**. Each
topic is one logical sheet (Guidelines, Run-to-Run Comparison, ERF
Table, …). The engineer picks which topics to include and what
annexure letter to use for each — both decisions live in the project
YAML under ``report.annexures`` and round-trip through the GUI.

This module owns:

  * The :class:`AnnexureTopic` dataclass.
  * Seven concrete topic definitions in :data:`TOPIC_REGISTRY`, in the
    canonical display order (A → G under the default lettering).
  * Helper :func:`default_annexure_selection` returning the legacy
    "E_F preset" equivalent used when a YAML has no
    ``report.annexures`` block (backward-compat).

The underlying sheet writers live in
:mod:`src.reports.annexure_writer` — the topic adapters here just
massage the ``AnalysisResult`` arguments into the shape each writer
needs and then call it. Adapters conform to a single uniform
signature::

    writer(workbook, sheet_name, project, results, run2_year) -> None

so the top-level builder can walk the topic list without per-topic
branching.
"""
from __future__ import annotations

from dataclasses import dataclass
from typing import Any, Callable, TYPE_CHECKING

if TYPE_CHECKING:
    from openpyxl.workbook import Workbook
    from src.models import Project


# ---------------------------------------------------------------------------
# Topic dataclass
# ---------------------------------------------------------------------------

@dataclass(frozen=True)
class AnnexureTopic:
    """One logical sheet in the report.

    Topics with ``implemented=False`` produce a placeholder sheet
    (currently only ``dent_strain_b318``). Selecting an unimplemented
    topic is a deliberate UI choice — the user gets a sheet that
    documents what's coming.
    """
    id: str
    display_name: str
    default_letter: str
    writer: Callable[..., None]
    implemented: bool = True


# ---------------------------------------------------------------------------
# Adapters: uniform (workbook, sheet_name, project, results, run2_year)
# ---------------------------------------------------------------------------
#
# Each adapter unpacks ``results`` (an :class:`AnalysisResult`) into the
# parameter shape the underlying writer needs. Writers themselves stay in
# ``src.reports.annexure_writer`` so this module doesn't accumulate Excel
# styling logic. All adapters are sync, side-effect-only (write into the
# workbook), and never raise on missing/empty inputs — they render a
# best-effort sheet so a partial AnalysisResult still produces output.


def _sorted_cgrs(results: Any) -> list:
    """Stable sort: ascending by run-2 abs_distance, tie-break on anomaly_id.

    Mirrors the AnnexureWriter sort in ``write()`` (annexure_writer.py
    line 171-174) so per-topic sheets agree on row order with each other.
    """
    cgrs = list(getattr(results, "cgr_results", []) or [])
    return sorted(
        cgrs,
        key=lambda r: (r.feature.abs_distance_m or 0.0, str(r.feature.anomaly_id)),
    )


def _topic_results_ili_comparison(
    workbook: "Workbook",
    sheet_name: str,
    project: "Project",
    results: Any,
    run2_year: int | None,
    *,
    title_text: str | None = None,
) -> None:
    """Wraps ``_write_annexure_e`` (run-to-run comparison)."""
    from src.reports.annexure_writer import AnnexureWriter
    writer = AnnexureWriter()
    year_new = str(run2_year) if run2_year else "run 2"
    year_old = (
        str(project.run_1.inspection_date.year)
        if (project.run_1 and project.run_1.inspection_date) else "run 1"
    )
    section_name = (
        project.pipeline.pipeline_name if project.pipeline else ""
    ) or ""
    writer._write_annexure_e(
        workbook, _sorted_cgrs(results),
        year_new=year_new, year_old=year_old, section_name=section_name,
        sheet_name=sheet_name, title_text=title_text,
    )


def _topic_metal_loss_anomalies(
    workbook: "Workbook",
    sheet_name: str,
    project: "Project",
    results: Any,
    run2_year: int | None,
    *,
    title_text: str | None = None,
) -> None:
    """Wraps ``_write_annexure_f`` (metal-loss anomalies + repair date)."""
    from src.reports.annexure_writer import AnnexureWriter
    writer = AnnexureWriter()
    ffp_by_id = {
        r.feature_id: r
        for r in (getattr(results, "ffp_results", []) or [])
    }
    pred_by_id = {
        p.feature_id: p
        for p in (getattr(results, "repair_predictions", []) or [])
    }
    pipeline = project.pipeline
    section_name = (pipeline.pipeline_name if pipeline else "") or ""
    pipeline_name = (
        project.project_name
        or (pipeline.client_name if pipeline else "")
        or section_name
    )
    writer._write_annexure_f(
        workbook, _sorted_cgrs(results), ffp_by_id, pred_by_id,
        pipeline=pipeline, section_name=section_name,
        pipeline_name=pipeline_name, project=project,
        sheet_name=sheet_name, title_text=title_text,
    )


def _topic_estimated_erf_defects(
    workbook: "Workbook",
    sheet_name: str,
    project: "Project",
    results: Any,
    run2_year: int | None,
    *,
    title_text: str | None = None,
) -> None:
    """Wraps ``_write_annexure_c`` (FFP at year 0 + +N-yr projection)."""
    from src.reports.annexure_writer import AnnexureWriter
    writer = AnnexureWriter()
    ffp_by_id = {
        r.feature_id: r
        for r in (getattr(results, "ffp_results", []) or [])
    }
    pred_by_id = {
        p.feature_id: p
        for p in (getattr(results, "repair_predictions", []) or [])
    }
    method_label = "ASME B-31G Original"
    cfg = project.config or {}
    primary = ((cfg.get("ffp") or {}).get("primary_method")) or ""
    if primary == "B31G_Modified":
        method_label = "ASME B-31G Modified"
    elif primary == "DNV_RP_F101":
        method_label = "DNV-RP-F101"
    elif primary == "RSTRENG":
        method_label = "RSTRENG"
    writer._write_bcd_assessment_sheet(
        workbook, sheet_name, method_label,
        _sorted_cgrs(results), ffp_by_id, pred_by_id,
        pipeline=project.pipeline, project=project,
        method_for_psafe=method_label,
        title_text=title_text,
    )


def _topic_estimated_erf_circ(
    workbook: "Workbook",
    sheet_name: str,
    project: "Project",
    results: Any,
    run2_year: int | None,
    *,
    title_text: str | None = None,
) -> None:
    """Build the Annexure-E equivalent (Kastner Psafe / ERF for every
    circumferential defect).

    v0.3.3 rewrite. The pre-v0.3.3 implementation filtered to features
    whose stored FFPResult had ``method is FFPMethod.KASTNER`` — i.e.
    only those where Kastner was the *controlling* method. For typical
    Indian metal-loss populations B31G's Psafe is lower than Kastner's
    net-section approximation, so the controlling method is almost
    never Kastner. The sheet came out empty for BPCL Mathura-Piyala
    (323 circ defects → 0 rows) despite the engine correctly computing
    Kastner for every one of them inside ``ffp_assess``.

    The published Athena reference deliverable convention is different:
    Annexure E (or D in the legacy preset) reports the Kastner Psafe /
    ERF for **every** circumferential defect, regardless of whether
    Kastner is the lower-Psafe / controlling method on that feature.
    The B31G column is reported separately in Annexure D.

    Implementation: walk the CGR list, filter via
    ``is_kastner_eligible`` (the multi-signal classifier added in
    v0.3.3 — POF enum / label substring / geometric proxy), and
    dispatch Kastner directly per feature using the same kwargs the
    coordinator uses. This bypasses ``ffps_by_id`` entirely — that
    dict only stores the controlling result per feature and so cannot
    surface Kastner data when B31G controls.
    """
    from src.reports.annexure_writer import AnnexureWriter
    from src.core.ffp import is_kastner_eligible, kastner
    from src.models import FFPMethod

    writer = AnnexureWriter()
    pipeline = project.pipeline
    pred_by_id = {
        p.feature_id: p
        for p in (getattr(results, "repair_predictions", []) or [])
    }
    cgrs_sorted = _sorted_cgrs(results)

    # Filter to circumferential defects via the v0.3.3 multi-signal
    # eligibility helper.
    circ_cgrs = [c for c in cgrs_sorted if is_kastner_eligible(c.feature)]

    # Dispatch Kastner per eligible feature. Skip rows where Kastner
    # can't run (missing width / WT / depth / MAOP zone) — those would
    # also have been silently skipped by the prior implementation.
    kastner_ffp_by_id: dict[str, Any] = {}
    for c in circ_cgrs:
        f = c.feature
        try:
            if f.width_mm is None or f.wt_mm is None \
                    or f.depth_pct_wt is None:
                continue
            zone, _zone_idx, _zone_fallback = pipeline.maop_for_feature(f)
            if zone is None:
                continue
            depth_mm = (f.depth_pct_wt / 100.0) * f.wt_mm
            r = kastner(
                d_mm=depth_mm,
                W_mm=f.width_mm,
                t_mm=f.wt_mm,
                D_mm=pipeline.diameter_mm,
                smys_mpa=pipeline.smys_mpa,
                Fd=zone.design_factor,
                maop_kgcm2=zone.maop_kgcm2,
                feature_id=f.anomaly_id,
            )
            kastner_ffp_by_id[f.anomaly_id] = r
        except (ValueError, AttributeError, TypeError, ZeroDivisionError):
            # Defensive — a bad row shouldn't kill the whole sheet.
            continue

    # Only emit rows that successfully produced a Kastner result.
    emit_cgrs = [c for c in circ_cgrs if c.feature.anomaly_id in kastner_ffp_by_id]

    writer._write_bcd_assessment_sheet(
        workbook, sheet_name, "Kastner Approach",
        emit_cgrs, kastner_ffp_by_id, pred_by_id,
        pipeline=pipeline, project=project,
        method_for_psafe="Kastner Approach",
        title_text=title_text,
    )


def _topic_qa_findings(
    workbook: "Workbook",
    sheet_name: str,
    project: "Project",
    results: Any,
    run2_year: int | None,
    *,
    title_text: str | None = None,
) -> None:
    """Wraps ``_write_issues_sheet`` (Quality Assurance findings)."""
    from src.reports.annexure_writer import AnnexureWriter
    flag_report = getattr(results, "flag_report", None)
    if flag_report is None:
        # No flag report -> nothing to write. Create an empty sheet
        # with just the title so the user sees the topic was selected
        # but produced no findings.
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        title_fill = PatternFill(start_color="FFC000", end_color="FFC000",
                                 fill_type="solid")
        ws = workbook.create_sheet(sheet_name)
        ws.cell(1, 1).value = title_text or sheet_name
        ws.cell(1, 1).font = Font(bold=True, size=12)
        ws.cell(1, 1).fill = title_fill
        ws.cell(1, 1).alignment = Alignment(horizontal="center")
        ws.cell(1, 1).border = border
        ws.merge_cells("A1:E1")
        ws.cell(2, 1).value = "QA: clean — no findings raised."
        return
    AnnexureWriter._write_issues_sheet(
        workbook, flag_report,
        sheet_name=sheet_name, title_text=title_text,
    )


def _topic_guidelines_formulas(
    workbook: "Workbook",
    sheet_name: str,
    project: "Project",
    results: Any,
    run2_year: int | None,
    *,
    title_text: str | None = None,
) -> None:
    """Wraps ``_write_guidelines_formulas_sheet`` (9-section guide)."""
    from src.reports.annexure_writer import _write_guidelines_formulas_sheet
    _write_guidelines_formulas_sheet(
        workbook, sheet_name, project, results, run2_year,
        title_text=title_text,
    )


def _topic_dent_strain_b318(
    workbook: "Workbook",
    sheet_name: str,
    project: "Project",
    results: Any,
    run2_year: int | None,
    *,
    title_text: str | None = None,
) -> None:
    """Wraps ``_write_dent_strain_sheet`` — v0.3.1 full implementation."""
    from src.reports.annexure_writer import _write_dent_strain_sheet
    _write_dent_strain_sheet(
        workbook, sheet_name, project, results, run2_year,
        title_text=title_text,
    )


# ---------------------------------------------------------------------------
# Registry
# ---------------------------------------------------------------------------

# Ordered list — the GUI displays topics in this order and the default
# annexure letters follow this sequence (A → G). User selection in
# ``report.annexures`` overrides both order and lettering.
_TOPIC_ORDER: list[AnnexureTopic] = [
    AnnexureTopic(
        id="guidelines_formulas",
        display_name="Guidelines & Formulas Used",
        default_letter="A",
        writer=_topic_guidelines_formulas,
    ),
    AnnexureTopic(
        id="results_ili_comparison",
        display_name="Results of ILI Comparison",
        default_letter="B",
        writer=_topic_results_ili_comparison,
    ),
    AnnexureTopic(
        id="metal_loss_anomalies",
        display_name="Metal Loss Anomalies with Repair Prediction",
        default_letter="C",
        writer=_topic_metal_loss_anomalies,
    ),
    AnnexureTopic(
        id="estimated_erf_defects",
        display_name="Estimated ERF of Defects (Year 0 + Future Projection)",
        default_letter="D",
        writer=_topic_estimated_erf_defects,
    ),
    AnnexureTopic(
        id="estimated_erf_circ",
        display_name="Estimated ERF of Circumferential Defects (Kastner)",
        default_letter="E",
        writer=_topic_estimated_erf_circ,
    ),
    AnnexureTopic(
        id="dent_strain_b318",
        display_name="Estimated Strain in Dents per ASME B31.8",
        default_letter="F",
        writer=_topic_dent_strain_b318,
        implemented=True,    # v0.3.1: full math (was placeholder in v0.2.5)
    ),
    AnnexureTopic(
        id="qa_findings",
        display_name="Quality Assurance Findings",
        default_letter="G",
        writer=_topic_qa_findings,
    ),
]

# Lookup by ID — used by YAML parser, GUI, builder.
TOPIC_REGISTRY: dict[str, AnnexureTopic] = {t.id: t for t in _TOPIC_ORDER}


def all_topics_in_order() -> list[AnnexureTopic]:
    """Return all registered topics in canonical display order.

    Returns a new list each call (callers can sort / filter freely).
    """
    return list(_TOPIC_ORDER)


# ---------------------------------------------------------------------------
# Backward-compat default selection
# ---------------------------------------------------------------------------

# When a YAML has no ``report.annexures`` block (pre-v0.2.5 projects),
# we fall back to the closest equivalent of the legacy "E_F" preset:
# the run-to-run comparison + metal-loss anomalies + QA sheet, with
# their default letters (B, C, G — not E, F, ... because v0.2.5
# renumbers the canonical letters; the *content* is the same).
_LEGACY_DEFAULT_TOPIC_IDS: tuple[str, ...] = (
    "results_ili_comparison",
    "metal_loss_anomalies",
    "qa_findings",
)


def default_annexure_selection() -> list[tuple[str, str]]:
    """Return the v0.2.0–v0.2.4 "E_F preset" equivalent as topic+letter pairs.

    Used when ``Project.from_yaml`` doesn't find a ``report.annexures``
    block, so existing customer YAMLs keep producing the same set of
    sheets without any migration step.

    Returned list is fresh on each call.
    """
    out: list[tuple[str, str]] = []
    for tid in _LEGACY_DEFAULT_TOPIC_IDS:
        t = TOPIC_REGISTRY[tid]
        out.append((t.id, t.default_letter))
    return out


__all__ = [
    "AnnexureTopic",
    "TOPIC_REGISTRY",
    "all_topics_in_order",
    "default_annexure_selection",
]
