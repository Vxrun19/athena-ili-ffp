"""
ILI pipe-tally reader.

Reads any vendor's Excel pipe tally (NGP single-sheet 2019, NGP multi-sheet
2023+, Athena 2018 Pipeline Tally + Metal Loss List, etc.) into an ILIRun.

>>> Add a new vendor format by adding column synonyms to
>>> /config/column_synonyms.yaml — no code changes needed.

ILI vendor files list both cluster parents and their child features.
For FFP assessment, use `ILIRun.features_for_assessment()`; for raw
feature counts (what the vendor PDF's "Total rows" figure shows) use
`ILIRun.features`. The reader populates `Feature.cluster_parent_id` on
clustered children using description markers from `column_synonyms.yaml`
(value_normalisations.cluster_child_markers), and `Feature.is_cluster_parent`
on rows whose feature_identification resolves to COCL.

The reader does five things, in order:

    1. Sheet discovery       — pick the sheet that actually holds defects.
    2. Header-row detection  — find the row of column names (vendors put
                                title/sub-title rows above the headers).
    3. Column mapping        — resolve each canonical field via the synonyms
                                config (case- and whitespace-insensitive).
    4. Row filtering         — drop welds/valves/supports/etc. via the
                                `skip` list; keep only metal-loss anomalies.
    5. Value normalisation   — every cell goes through parsers in
                                `src.models.units` (parse_clock, parse_surface,
                                parse_depth). Vendor codes like "MILL", "INT"
                                are mapped via `value_normalisations` in
                                column_synonyms.yaml BEFORE the parser sees
                                them.

The resulting ILIRun has a full audit trail (sheet picked, header row index,
column map used, rows-filtered-by-reason) so downstream QA and the final
report can show "why this row didn't make it in."
"""
from __future__ import annotations

import bisect
import re
import unicodedata
from collections import Counter
from pathlib import Path
from typing import Any

import yaml

from src.models import (
    DimensionClass,
    Feature,
    FeatureIdentification,
    ILIRun,
    Joint,
)
from src.models.units import parse_clock, parse_depth, parse_surface
from src.validation import QAFlag, QAFlagCode, QASeverity, make_flag

# fids that are NOT metal-loss anomalies — they may share the "Anomaly"
# feature-type label, but B31G/RSTRENG don't apply to them. Dents, weld
# anomalies, and cracks are assessed by separate methods, not this tool.
_NON_METAL_LOSS_FIDS: frozenset[FeatureIdentification] = frozenset(
    {
        FeatureIdentification.GIRTH_WELD_ANOMALY,
        FeatureIdentification.SPIRAL_WELD_ANOMALY,
        FeatureIdentification.LONG_WELD_ANOMALY,
        FeatureIdentification.DENT,
        FeatureIdentification.DENT_WITH_METAL_LOSS,
        FeatureIdentification.CRACK,
        FeatureIdentification.RIPPLE,
    }
)

# Default coordinate bounding box — Indian subcontinent. Mirrors
# qa.coordinate_bounds in config/default_project.yaml; keep in sync.
_DEFAULT_LAT_BOUNDS: tuple[float, float] = (6.0, 38.0)
_DEFAULT_LON_BOUNDS: tuple[float, float] = (68.0, 98.0)

# A backward jump larger than this in the abs_distance sequence flips the
# reader into chainage-lookup mode (rows aren't reliably chainage-ordered,
# so forward-fill of joint context would mis-attribute features). Tuned
# above any plausible within-joint reordering (~12 m joints) but well
# below realistic file-corruption-scale jumps. 500 m is generous.
_NON_MONOTONIC_JUMP_M: float = 500.0


def _prescan_chainage(
    rows: list[list[Any]],
    header_idx: int,
    col_map: dict[str, int],
) -> tuple[bool, list[float], list[int]]:
    """Single pass over rows. Detects chainage non-monotonicity and collects
    sorted weld anchors for chainage-lookup mode.

    Returns (is_monotonic, weld_xs, weld_jnos) where weld_xs/jnos are paired
    parallel lists sorted ascending by abs_distance.

    Non-monotonic detection: any backward jump > `_NON_MONOTONIC_JUMP_M`
    flips the flag. Within-joint reorderings (a few metres) don't count.

    Anchor collection: every row that has BOTH an explicit joint_number
    and a valid abs_distance contributes one anchor. In NGP 2019 single-
    sheet only weld rows carry an explicit jno (anomalies have jno=None
    and inherit from the preceding weld), so anchors == welds there. In
    flat formats every row contributes but we only use the anchor list
    when non-monotonic, which those files aren't.
    """
    is_monotonic = True
    prev_abs: float | None = None
    anchors: list[tuple[float, int]] = []
    for row in rows[header_idx + 1 :]:
        jno = _to_int(_get(row, col_map, "joint_number"))
        abs_d = _to_float(_get(row, col_map, "abs_distance_m"))
        if jno is not None and abs_d is not None:
            anchors.append((abs_d, jno))
        if abs_d is not None:
            if prev_abs is not None and abs_d < prev_abs - _NON_MONOTONIC_JUMP_M:
                is_monotonic = False
            prev_abs = abs_d
    anchors.sort()
    xs = [a[0] for a in anchors]
    jnos = [a[1] for a in anchors]
    return is_monotonic, xs, jnos


def _chainage_lookup_joint(
    abs_d: float | None,
    weld_xs: list[float],
    weld_jnos: list[int],
) -> tuple[int | None, float | None]:
    """Binary-search for the joint containing this abs_distance.

    Returns (joint_number, upstream_weld_dist_m). The recomputed
    upstream_weld_dist_m is the feature's position relative to the joint's
    upstream weld (always non-negative in the standard convention).

    If abs_distance falls before any known weld, returns (None, None).
    """
    if abs_d is None or not weld_xs:
        return None, None
    # bisect_right with abs_d gives the first index AFTER abs_d; we want
    # the largest weld at or before abs_d, so subtract one.
    i = bisect.bisect_right(weld_xs, abs_d)
    if i == 0:
        return None, None
    joint = weld_jnos[i - 1]
    uw = float(abs_d) - float(weld_xs[i - 1])
    return joint, uw

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

_PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_SYNONYMS_PATH = _PROJECT_ROOT / "config" / "column_synonyms.yaml"

# Canonical field set the reader knows about.
_REQUIRED_FIELDS = (
    "abs_distance_m",
    "joint_number",
    "depth_pct_wt",
    "length_mm",
    "width_mm",
    "surface",
    "wt_mm",
)
_OPTIONAL_FIELDS = (
    "clock_position",
    "feature_type",
    "feature_identification",
    "dimension_class",
    "description",
    "anomaly_id",
    "erf",
    "psafe",
    "latitude",
    "longitude",
    "altitude_m",
    "upstream_weld_dist_m",
    "joint_length_m",
)
_ALL_FIELDS = _REQUIRED_FIELDS + _OPTIONAL_FIELDS

# Min canonical-field hits to call a sheet a "defect sheet" or a row a header.
_DEFECT_SHEET_MIN_HITS = 4
_HEADER_ROW_MIN_HITS = 4

# Sheet names we prefer outright (lowercased) when present.
_PREFERRED_SHEET_NAMES = ("defects", "metal loss list", "severity list")
_NEG_SHEET_NAMES = (
    "weld",
    "casing",
    "wall thickness",
    "reference point",
    "bend",
    "installation",
    "pipe",
    "adjacent",
)

_NA_STRINGS = {"", "-", "--", "–", "—", "n/a", "na", "·", ".", "?", "not defined"}


# ---------------------------------------------------------------------------
# Synonym handling
# ---------------------------------------------------------------------------

_NORM_RE = re.compile(r"[\s_\-/.,()\[\]{}'\"`²°*:%]+")


def _norm(s: Any) -> str:
    """Normalise a column header or value for comparison.

    Lowercase, strip whitespace/newlines, replace punctuation runs with single
    space, collapse, NFKC-normalise. Tolerates trailing/leading junk.
    """
    if s is None:
        return ""
    if not isinstance(s, str):
        s = str(s)
    s = unicodedata.normalize("NFKC", s)
    s = s.replace("\n", " ").replace("\r", " ").replace("\t", " ")
    s = s.casefold().strip()
    s = _NORM_RE.sub(" ", s).strip()
    return s


def load_synonyms(path: str | Path | None = None) -> dict[str, Any]:
    """Load the column-synonyms YAML."""
    path = Path(path) if path else DEFAULT_SYNONYMS_PATH
    with path.open("r", encoding="utf-8") as f:
        data = yaml.safe_load(f) or {}
    return data


def _build_synonym_index(synonyms: dict[str, Any]) -> dict[str, str]:
    """Build a single normalised-header -> canonical-field lookup.

    Order in the YAML defines precedence: if two canonical fields list the
    same synonym, the one declared first wins (mirrors the YAML comment
    "Order matters: more specific patterns first").
    """
    idx: dict[str, str] = {}
    for canonical in _ALL_FIELDS:
        entry = synonyms.get(canonical) or {}
        for syn in entry.get("synonyms", []) or []:
            n = _norm(syn)
            if n and n not in idx:
                idx[n] = canonical
    return idx


def _build_value_norm_index(synonyms: dict[str, Any]) -> dict[str, dict[str, str]]:
    """For each value-normalisation group, map normalised raw -> canonical.

    Groups shaped as `{canonical: [raws...]}` produce a raw->canonical lookup.
    Groups shaped as a plain list (e.g. `cluster_child_markers: [...]`) are
    consumed elsewhere and skipped here.
    """
    vn = synonyms.get("value_normalisations") or {}
    out: dict[str, dict[str, str]] = {}
    for group, mapping in vn.items():
        if not isinstance(mapping, dict):
            continue
        bucket: dict[str, str] = {}
        for canonical, raws in mapping.items():
            for r in raws or []:
                bucket[_norm(r)] = canonical
        out[group] = bucket
    return out


def _normalise_value(group: str, raw: Any, vn_index: dict[str, dict[str, str]]) -> str | None:
    """Look up a free-text value in the value-normalisations table."""
    table = vn_index.get(group) or {}
    return table.get(_norm(raw))


# ---------------------------------------------------------------------------
# Workbook abstraction (xlsx via openpyxl; xls via xlrd; detect by magic)
# ---------------------------------------------------------------------------

_XLSX_MAGIC = b"PK\x03\x04"
_XLS_MAGIC = b"\xd0\xcf\x11\xe0"


def _detect_format(path: Path) -> str:
    with path.open("rb") as f:
        head = f.read(8)
    if head.startswith(_XLSX_MAGIC):
        return "xlsx"
    if head.startswith(_XLS_MAGIC):
        return "xls"
    raise ValueError(
        f"{path.name}: unrecognised file format (magic={head!r}). "
        "Expected .xlsx (PK..) or .xls (D0CF11E0..)."
    )


def _load_sheets(path: Path) -> dict[str, list[list[Any]]]:
    """Return {sheet_name: rows-as-lists-of-cell-values}.

    Done eagerly because we need to score sheets and then re-read the chosen
    one anyway. For very large files this still streams in O(rows) memory.
    """
    fmt = _detect_format(path)
    if fmt == "xlsx":
        return _load_xlsx_sheets(path)
    return _load_xls_sheets(path)


def _load_xlsx_sheets(path: Path) -> dict[str, list[list[Any]]]:
    import openpyxl
    import warnings as _warnings

    with _warnings.catch_warnings():
        # Suppress benign "Print area cannot be set to Defined name" etc.
        _warnings.simplefilter("ignore", category=UserWarning)
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        out: dict[str, list[list[Any]]] = {}
        for sn in wb.sheetnames:
            ws = wb[sn]
            out[sn] = [list(row) for row in ws.iter_rows(values_only=True)]
        return out
    finally:
        wb.close()


def _load_xls_sheets(path: Path) -> dict[str, list[list[Any]]]:
    import xlrd  # type: ignore[import-not-found]
    book = xlrd.open_workbook(str(path))
    out: dict[str, list[list[Any]]] = {}
    for sn in book.sheet_names():
        sh = book.sheet_by_name(sn)
        rows: list[list[Any]] = []
        for r in range(sh.nrows):
            rows.append([sh.cell_value(r, c) for c in range(sh.ncols)])
        out[sn] = rows
    return out


# ---------------------------------------------------------------------------
# Sheet & header detection
# ---------------------------------------------------------------------------

def _score_row_for_headers(row: list[Any], syn_idx: dict[str, str]) -> tuple[int, set[str]]:
    hits: set[str] = set()
    for cell in row:
        canonical = syn_idx.get(_norm(cell))
        if canonical:
            hits.add(canonical)
    return len(hits), hits


def _find_header_row(rows: list[list[Any]], syn_idx: dict[str, str]) -> tuple[int, set[str]]:
    """Scan the first 10 rows; return (idx, hit-set) of the row with most hits."""
    best_idx = 0
    best_hits: set[str] = set()
    for i, row in enumerate(rows[:10]):
        n, hits = _score_row_for_headers(row, syn_idx)
        if n > len(best_hits):
            best_idx = i
            best_hits = hits
    return best_idx, best_hits


def _pick_defect_sheet(
    sheets: dict[str, list[list[Any]]],
    syn_idx: dict[str, str],
) -> tuple[str, int, set[str]]:
    """Choose the sheet most likely to hold defects. Returns (sheet, header_idx, hits)."""
    # Pass 1: any sheet whose name matches the preferred list AND scores well.
    candidates: list[tuple[str, int, set[str], int]] = []
    for sn, rows in sheets.items():
        if not rows:
            continue
        idx, hits = _find_header_row(rows, syn_idx)
        if len(hits) < _DEFECT_SHEET_MIN_HITS:
            continue
        sn_low = sn.casefold()
        is_preferred = any(p in sn_low for p in _PREFERRED_SHEET_NAMES)
        is_negative = any(p in sn_low for p in _NEG_SHEET_NAMES)
        priority = 2 if is_preferred else (0 if is_negative else 1)
        candidates.append((sn, idx, hits, priority))

    if not candidates:
        # No sheet scored at all — diagnostic for the caller.
        diag = ", ".join(f"{sn!r}({len(rows)} rows)" for sn, rows in sheets.items())
        raise ValueError(
            f"No sheet contains enough canonical defect columns "
            f"(need ≥{_DEFECT_SHEET_MIN_HITS} of {_ALL_FIELDS}). "
            f"Sheets seen: {diag}"
        )

    # Sort by (priority desc, hit count desc, sheet-name).
    candidates.sort(key=lambda c: (-c[3], -len(c[2]), c[0]))
    sn, idx, hits, _ = candidates[0]
    return sn, idx, hits


# ---------------------------------------------------------------------------
# Column map
# ---------------------------------------------------------------------------

def _build_column_map(
    header_row: list[Any],
    syn_idx: dict[str, str],
) -> dict[str, int]:
    """Map canonical-field -> column-index. First match wins."""
    mapping: dict[str, int] = {}
    for j, cell in enumerate(header_row):
        canonical = syn_idx.get(_norm(cell))
        if canonical and canonical not in mapping:
            mapping[canonical] = j
    return mapping


def _check_required(column_map: dict[str, int], header_row: list[Any], syn_idx: dict[str, str]) -> None:
    missing = [f for f in _REQUIRED_FIELDS if f not in column_map]
    if not missing:
        return
    seen = [str(c) for c in header_row if c not in (None, "")]
    tried = sorted({k for k, v in syn_idx.items() if v in missing})
    raise ValueError(
        f"Required columns missing: {missing}. "
        f"Headers seen: {seen}. "
        f"Synonyms tried for missing fields: {tried[:30]}{'…' if len(tried) > 30 else ''}"
    )


# ---------------------------------------------------------------------------
# Cell helpers
# ---------------------------------------------------------------------------

def _get(row: list[Any], col_map: dict[str, int], field_name: str) -> Any:
    j = col_map.get(field_name)
    if j is None or j >= len(row):
        return None
    return row[j]


def _is_na(v: Any) -> bool:
    if v is None:
        return True
    if isinstance(v, str):
        return _norm(v) in _NA_STRINGS or v.strip() == ""
    return False


def _to_float(v: Any) -> float | None:
    if _is_na(v):
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip()
        # European comma decimals (50,5 -> 50.5) — only if no period present.
        if "," in s and "." not in s:
            s = s.replace(",", ".")
        # Unicode minus, en-dash etc.
        s = s.replace("−", "-").replace("–", "-")
        try:
            return float(s)
        except ValueError:
            return None
    return None


def _to_int(v: Any) -> int | None:
    f = _to_float(v)
    if f is None:
        return None
    try:
        return int(round(f))
    except (ValueError, OverflowError):
        return None


def _to_str(v: Any) -> str | None:
    if _is_na(v):
        return None
    if isinstance(v, str):
        return v.strip() or None
    return str(v)


# ---------------------------------------------------------------------------
# Skip / keep logic
# ---------------------------------------------------------------------------

def _row_is_anomaly(
    row: list[Any],
    col_map: dict[str, int],
    skip_set: set[str],
    keep_set: set[str],
    vn_fid: dict[str, str],
    has_depth: bool,
) -> tuple[bool, str]:
    """Decide whether a row is a metal-loss defect. Returns (keep, reason_if_skip).

    Logic, in order:
      1. Explicit skip-list match on ftype OR fid → drop.
      2. Explicit keep-list match → keep.
      3. fid resolves to a known POF code via value_normalisations → keep
         (post-parse `_NON_METAL_LOSS_FIDS` drop handles GWAN/DENT/etc).
      4. No type info at all → fall back on "has depth" heuristic.
      5. Type present but unrecognised → drop.
    """
    ftype = _to_str(_get(row, col_map, "feature_type"))
    fid = _to_str(_get(row, col_map, "feature_identification"))

    # 1. skip-list
    for candidate in (ftype, fid):
        if candidate is None:
            continue
        if _norm(candidate) in skip_set:
            return False, f"skip_list:{candidate}"

    # 2. keep-list (text)
    for candidate in (ftype, fid):
        if candidate is None:
            continue
        if _norm(candidate) in keep_set:
            return True, ""

    # 3. fid resolves to a known POF code. Whether it's anomaly-shaped
    #    (CORR/COCL) or weld-shaped (GWAN) the post-parse drop sorts it out;
    #    here we just refuse to filter rows whose fid maps to a real code.
    if fid is not None and _norm(fid) in vn_fid:
        return True, ""

    # 4. no type info
    if ftype is None and fid is None:
        return (True if has_depth else False), ("no_type_no_depth" if not has_depth else "")

    # 5. type present but unrecognised
    return False, f"unrecognised_type:{ftype or fid}"


# ---------------------------------------------------------------------------
# Coordinate sanity (for Indian pipelines)
# ---------------------------------------------------------------------------

def _resolve_coord_bounds(
    spec: dict | tuple | None,
) -> tuple[tuple[float, float], tuple[float, float]]:
    """Normalise a coordinate-bounds spec into (lat_bounds, lon_bounds)."""
    if spec is None:
        return _DEFAULT_LAT_BOUNDS, _DEFAULT_LON_BOUNDS
    if isinstance(spec, dict):
        lat = spec.get("lat") or list(_DEFAULT_LAT_BOUNDS)
        lon = spec.get("lon") or list(_DEFAULT_LON_BOUNDS)
        return (float(lat[0]), float(lat[1])), (float(lon[0]), float(lon[1]))
    if isinstance(spec, (tuple, list)) and len(spec) == 4:
        a, b, c, d = spec
        return (float(a), float(b)), (float(c), float(d))
    raise ValueError(
        f"coordinate_bounds must be dict {{'lat':[lo,hi],'lon':[lo,hi]}} "
        f"or 4-tuple (lat_lo,lat_hi,lon_lo,lon_hi); got {spec!r}"
    )


def _median(vals: list[float]) -> float:
    s = sorted(vals)
    n = len(s)
    return s[n // 2] if n % 2 else (s[n // 2 - 1] + s[n // 2]) / 2.0


def _check_coordinate_bounds(
    features: list[Feature],
    lat_bounds: tuple[float, float],
    lon_bounds: tuple[float, float],
) -> list[QAFlag]:
    """Sanity-check feature coordinates against an expected bounding box.

    Uses medians (robust to a handful of outlier rows). Three outcomes:
      1. Medians fall inside the expected bands -> no flag.
      2. Medians fall inside the SWAPPED bands -> swap every feature's
         (latitude, longitude) and emit COORDINATES_SWAPPED.
      3. Neither orientation puts both medians in band -> leave values as-is
         and emit LAT_LON_OUT_OF_BOUNDS (caller decides how loud to be).
    """
    lats = [f.latitude for f in features if f.latitude is not None]
    lons = [f.longitude for f in features if f.longitude is not None]
    if not lats or not lons:
        return []

    lat_med = _median(lats)
    lon_med = _median(lons)

    def _in(v: float, lo_hi: tuple[float, float]) -> bool:
        return lo_hi[0] <= v <= lo_hi[1]

    if _in(lat_med, lat_bounds) and _in(lon_med, lon_bounds):
        return []

    if _in(lon_med, lat_bounds) and _in(lat_med, lon_bounds):
        for f in features:
            f.latitude, f.longitude = f.longitude, f.latitude
        return [
            QAFlag(
                code=QAFlagCode.COORDINATES_SWAPPED,
                severity=QASeverity.WARN,
                message=(
                    f"Latitude/longitude appear swapped at source "
                    f"(median lat={lat_med:.3f}, lon={lon_med:.3f} fits lat-bounds"
                    f" {lon_bounds} and lon-bounds {lat_bounds}); values auto-swapped."
                ),
                context={
                    "lat_median": lat_med,
                    "lon_median": lon_med,
                    "lat_bounds": list(lat_bounds),
                    "lon_bounds": list(lon_bounds),
                },
            )
        ]

    return [
        QAFlag(
            code=QAFlagCode.LAT_LON_OUT_OF_BOUNDS,
            severity=QASeverity.ERROR,
            message=(
                f"Coordinates out of expected bounds in either orientation: "
                f"median lat={lat_med:.3f} (expected {lat_bounds}), "
                f"median lon={lon_med:.3f} (expected {lon_bounds}). "
                "Verify the file's CRS or override qa.coordinate_bounds."
            ),
            context={
                "lat_median": lat_med,
                "lon_median": lon_med,
                "lat_bounds": list(lat_bounds),
                "lon_bounds": list(lon_bounds),
            },
        )
    ]


# ---------------------------------------------------------------------------
# The reader itself
# ---------------------------------------------------------------------------

class ILIReader:
    """Config-driven reader for vendor pipe-tally Excel files.

    Usage:
        reader = ILIReader()                 # uses config/column_synonyms.yaml
        run = reader.read("tally.xlsx", run_id="run_1")

    Override the coordinate bounding box per project by passing
    `coordinate_bounds={"lat": [lo, hi], "lon": [lo, hi]}` or a 4-tuple
    `(lat_lo, lat_hi, lon_lo, lon_hi)`. Defaults match
    `qa.coordinate_bounds` in `config/default_project.yaml` (Indian
    subcontinent).
    """

    def __init__(
        self,
        synonyms_path: str | Path | None = None,
        coordinate_bounds: dict | tuple | None = None,
    ):
        self.synonyms = load_synonyms(synonyms_path)
        self._syn_idx = _build_synonym_index(self.synonyms)
        self._vn_idx = _build_value_norm_index(self.synonyms)
        self._skip_set: set[str] = set()
        self._keep_set: set[str] = set()
        ft_anom = (self.synonyms.get("value_normalisations") or {}).get(
            "feature_type_anomaly"
        ) or {}
        for s in ft_anom.get("skip", []) or []:
            self._skip_set.add(_norm(s))
        for s in ft_anom.get("is_anomaly", []) or []:
            self._keep_set.add(_norm(s))

        # Cluster-child markers from value_normalisations.cluster_child_markers.
        self._cluster_child_markers: tuple[str, ...] = tuple(
            _norm(m)
            for m in (self.synonyms.get("value_normalisations") or {}).get(
                "cluster_child_markers", []
            )
            or []
            if m
        )

        # Coordinate bounding box.
        self._lat_bounds, self._lon_bounds = _resolve_coord_bounds(coordinate_bounds)

        # Per-read scratch — collects vendor fid strings that LOOK
        # non-metal-loss but didn't normalise to a known POF code. Each
        # read() flushes this into the run's parse_warnings so a
        # maintainer can extend column_synonyms.yaml if a new vendor
        # variant shows up.
        self._unresolved_non_ml_strings: set[str] = set()

    # ------------------------------------------------------------------

    def read(
        self,
        filepath: str | Path,
        run_id: str = "run",
    ) -> ILIRun:
        path = Path(filepath)
        if not path.exists():
            raise FileNotFoundError(f"{path}")

        # Reset per-read scratch.
        self._unresolved_non_ml_strings = set()

        sheets = _load_sheets(path)
        sheet_name, header_idx, _hits = _pick_defect_sheet(sheets, self._syn_idx)
        rows = sheets[sheet_name]
        header_row = rows[header_idx]

        col_map = _build_column_map(header_row, self._syn_idx)
        _check_required(col_map, header_row, self._syn_idx)

        run = ILIRun(
            run_id=run_id,
            file_path=str(path),
            sheet_name=sheet_name,
            header_row_idx=header_idx,
            column_map=dict(col_map),
        )

        filtered = Counter()
        features: list[Feature] = []

        # Pre-scan: detect non-monotonic chainage + collect weld anchors for
        # lookup mode. For monotonic files this is cheap and the anchors go
        # unused; for non-monotonic files (HMEL run 1's anomaly-block layout)
        # it's how anomaly rows find their real joint.
        is_monotonic, weld_xs, weld_jnos = _prescan_chainage(
            rows, header_idx, col_map
        )
        use_chainage_lookup = (not is_monotonic) and bool(weld_xs)
        if use_chainage_lookup:
            run.parse_warnings.append("RECONSTRUCTED_JOINT_CONTEXT_FROM_CHAINAGE")
            run.qa_flags.append(make_flag(
                QAFlagCode.RECONSTRUCTED_JOINT_CONTEXT_FROM_CHAINAGE,
                "Row sequence is not chainage-monotonic — joint attribution "
                "for anomaly rows uses binary search against the sorted "
                "weld-anchor list rather than forward-fill of context.",
                context={"weld_anchor_count": len(weld_xs)},
            ))

        # Forward-filled joint context for hierarchical layouts (NGP 2019
        # single-sheet: welds carry j.no. + j.length, anomalies inherit them
        # from the preceding weld row). For flat layouts (every row has j.no.)
        # this is a no-op. For non-monotonic files we still maintain ctx_joint
        # for joint_length tracking via welds, but per-anomaly joint
        # attribution switches to chainage lookup below.
        ctx_joint: int | None = None

        # Full joint registry, populated from EVERY row with an explicit
        # joint number — welds, valves, supports, etc. — not just kept
        # features. This is what makes downstream joint alignment possible:
        # the matcher needs the complete joint sequence, not just joints
        # that happen to have defects.
        joints_data: dict[int, dict[str, Any]] = {}

        # Cluster context: when a COCL parent row is seen, subsequent CORR
        # rows whose description is marked as "grouped" / "child" are tagged
        # as children of that parent. Resets at joint boundaries so a parent
        # in one joint can't claim children in the next.
        ctx_cluster_parent: str | None = None
        ctx_cluster_joint: int | None = None

        for i, row in enumerate(rows[header_idx + 1 :], start=header_idx + 1):
            run.rows_read += 1

            # Quick skip: entirely blank rows.
            if not any(v not in (None, "") for v in row):
                filtered["blank_row"] += 1
                continue

            # Update forward-fill context BEFORE filtering — even skipped rows
            # (welds, supports) supply joint context for the rows that follow.
            jno_here = _to_int(_get(row, col_map, "joint_number"))
            if jno_here is not None:
                ctx_joint = jno_here
                entry = joints_data.setdefault(
                    jno_here, {"starts": [], "wts": [], "length": None}
                )
                # abs_distance on a row in the joint — for weld rows on a
                # multi-row format this is the joint start; for defect rows
                # it's the defect position (subtract upstream-weld dist).
                abs_d = _to_float(_get(row, col_map, "abs_distance_m"))
                if abs_d is not None:
                    uw = _to_float(_get(row, col_map, "upstream_weld_dist_m"))
                    entry["starts"].append(abs_d - uw if uw is not None else abs_d)
                wt_here = _to_float(_get(row, col_map, "wt_mm"))
                if wt_here is not None:
                    entry["wts"].append(wt_here)
            jlen_here = _to_float(_get(row, col_map, "joint_length_m"))
            if jlen_here is not None and ctx_joint is not None:
                entry = joints_data.setdefault(
                    ctx_joint, {"starts": [], "wts": [], "length": None}
                )
                if entry["length"] is None:
                    entry["length"] = jlen_here

            has_depth_raw = not _is_na(_get(row, col_map, "depth_pct_wt"))
            keep, reason = _row_is_anomaly(
                row,
                col_map,
                self._skip_set,
                self._keep_set,
                self._vn_idx.get("feature_identification", {}),
                has_depth_raw,
            )
            if not keep:
                filtered[reason or "filtered"] += 1
                # If the row was dropped because its fid LOOKS dent-like
                # / weld-like / crack-like but isn't in the value-
                # normalisation map, surface that loudly so a maintainer
                # extends column_synonyms.yaml. This is the Abu Road
                # 1ZYC failure mode in disguise — the row IS being
                # dropped (good!) but the operator gets no warning that
                # a vendor variant slipped past the map.
                if reason.startswith("unrecognised_type:"):
                    label = reason.split(":", 1)[1]
                    label_lo = label.strip().lower()
                    if label_lo and any(
                        tok in label_lo
                        for tok in ("dent", "weld", "crack", "lwcr",
                                    "mian", "miac")
                    ):
                        self._unresolved_non_ml_strings.add(label.strip())
                continue

            uw_override: float | None = None
            if jno_here is not None:
                joint = jno_here
            elif use_chainage_lookup:
                # Forward-fill is unreliable here (rows aren't chainage-
                # monotonic). Find the right joint by binary-searching the
                # weld anchors against this row's abs_distance, and
                # recompute upstream_weld_dist_m so it's the standard
                # non-negative "distance from the joint's upstream weld."
                row_abs_d = _to_float(_get(row, col_map, "abs_distance_m"))
                joint, uw_override = _chainage_lookup_joint(
                    row_abs_d, weld_xs, weld_jnos
                )
            else:
                joint = ctx_joint
            if joint is None:
                filtered["missing_joint_number"] += 1
                continue

            try:
                feat = self._row_to_feature(
                    row, col_map, run_id, i, joint, uw_override=uw_override
                )
            except ValueError as e:
                # Feature __post_init__ rejected the row — capture and skip.
                run.parse_warnings.append(f"row {i}: {e}")
                filtered["invalid_values"] += 1
                continue

            # Post-parse drops:
            #   1. (PREVIOUSLY) non-metal-loss fids were dropped here.
            #      That filter now lives inside
            #      ILIRun.features_for_assessment() so dents / welds /
            #      cracks remain visible in `run.features` (the raw,
            #      all-feature list) — useful for cross-checking the
            #      vendor's "Total rows" figure. Downstream tools
            #      (defect matcher, FFP coordinator, report writer)
            #      already iterate features_for_assessment() and so
            #      transparently exclude them.
            #
            #      The reader still records the filter in `filtered`
            #      so the rows_filtered report stays consistent, but
            #      the row is KEPT in features.
            if feat.feature_identification in _NON_METAL_LOSS_FIDS:
                filtered[f"non_metal_loss_fid:{feat.feature_identification.value}"] += 1
                # NOTE: do NOT `continue` here — keep the row in
                # `features`. The exclusion happens at consumption
                # time via features_for_assessment().
            #   2. No depth measurement — defect can't be FFP-assessed.
            if feat.depth_pct_wt is None:
                filtered["no_depth"] += 1
                continue

            # Cluster bookkeeping. Reset context if the joint changed.
            if ctx_cluster_joint != feat.joint_number:
                ctx_cluster_parent = None
                ctx_cluster_joint = feat.joint_number
            if feat.is_cluster_parent:
                # Become the new parent for subsequent rows in this joint.
                ctx_cluster_parent = feat.anomaly_id
            elif (
                ctx_cluster_parent is not None
                and self._cluster_child_markers
                and self._description_marks_child(feat.raw_description)
            ):
                feat.cluster_parent_id = ctx_cluster_parent

            features.append(feat)

        run.features = features
        run.rows_filtered = dict(filtered)
        run.skipped_count = sum(filtered.values())

        run.qa_flags.extend(
            _check_coordinate_bounds(features, self._lat_bounds, self._lon_bounds)
        )
        # Mirror QA flag text into parse_warnings for backwards-compat with
        # any caller already reading that list (the full QA pipeline in
        # Prompt 8 will deprecate parse_warnings).
        for flag in run.qa_flags:
            run.parse_warnings.append(str(flag))

        # Flush any unresolved dent-like / weld-like fid strings so a
        # maintainer notices and can extend value_normalisations in
        # column_synonyms.yaml. CRUCIALLY these rows are NOT
        # auto-filtered — the post-parse non-ML drop tests for explicit
        # enum membership and UNDEFINED is not in the set. The string
        # likely got swallowed into FFP with a bogus Psafe value (see
        # the Abu Road #1637 dent bug). Surface it loudly.
        for unresolved in sorted(self._unresolved_non_ml_strings):
            run.parse_warnings.append(
                f"UNRECOGNISED_NON_ML_FID: vendor sent {unresolved!r} but "
                "value_normalisations.feature_identification doesn't "
                "map it. Row was kept as UNDEFINED and may have leaked "
                "into FFP assessment — extend config/column_synonyms.yaml "
                "if this is a new dent/weld/crack variant."
            )

        # Merge joint data from a secondary "Pipe" / "Pipeline Tally" sheet
        # if present. Multi-sheet files (NGP 2023+, Athena 2018) keep the
        # full joint list there; the Defects/Metal-Loss-List sheet only has
        # defect-bearing joints.
        self._merge_joints_from_secondary_sheets(sheets, sheet_name, joints_data)

        run.joints = self._build_joints(joints_data)
        return run

    # ------------------------------------------------------------------

    def _description_marks_child(self, description: str) -> bool:
        if not description:
            return False
        n = _norm(description)
        return any(m in n for m in self._cluster_child_markers)

    # ------------------------------------------------------------------

    def _row_to_feature(
        self,
        row: list[Any],
        col_map: dict[str, int],
        run_id: str,
        row_idx: int,
        joint_number: int,
        uw_override: float | None = None,
    ) -> Feature:
        wt = _to_float(_get(row, col_map, "wt_mm"))
        # Depth column convention note: NGP/Athena tally files use a
        # single "Depth, %WT/OD" header — the value is interpreted as
        # %WT for metal-loss rows but as %OD for dents (a dent at 0.9
        # means 0.9% OD, not 90% WT). We don't fork the conversion
        # here because dents are filtered out below via
        # _NON_METAL_LOSS_FIDS BEFORE they reach the FFP engine. If a
        # future maintainer ever adds dent-aware analysis, this is the
        # spot to branch on feature_identification and convert via the
        # OD/WT ratio. Until then the "treat all values as %WT"
        # assumption is safe — but only for the rows that survive the
        # non-ML drop. The 1ZYC Abu Road bug (#125-equivalent for
        # dents) traces back to a dent string that slipped past
        # normalization and was assessed with %OD numbers treated as
        # %WT, producing ERF ≈ 8.57. Bottom line: keep the dent filter
        # tight.
        depth_raw = _get(row, col_map, "depth_pct_wt")
        depth_pct, _depth_mm = parse_depth(depth_raw, wt)

        clock = parse_clock(_get(row, col_map, "clock_position"))
        surface = parse_surface(_get(row, col_map, "surface"))

        # Map vendor free-text to canonical POF codes via value_normalisations.
        fid_raw = _get(row, col_map, "feature_identification")
        fid_code = (
            _normalise_value("feature_identification", fid_raw, self._vn_idx)
            if fid_raw is not None
            else None
        )
        try:
            fid_enum = FeatureIdentification(fid_code) if fid_code else FeatureIdentification.UNDEFINED
        except ValueError:
            fid_enum = FeatureIdentification.UNDEFINED

        # Parse warning: vendor sent a non-empty string that didn't
        # resolve to a known POF code (UNDEFINED). For most metal-loss
        # variants this is harmless, but for strings that LOOK dent-
        # like or weld-like the row will silently slip past the
        # non-metal-loss filter further down and end up in FFP — which
        # is the exact 1ZYC Abu Road bug. Surface the leak so a
        # maintainer can add the missing string to value_normalisations
        # in column_synonyms.yaml.
        if fid_enum is FeatureIdentification.UNDEFINED and fid_raw:
            raw_lo = str(fid_raw).strip().lower()
            if raw_lo and any(
                tok in raw_lo
                for tok in ("dent", "weld", "crack", "lwcr", "mian", "miac")
            ):
                self._unresolved_non_ml_strings.add(str(fid_raw).strip())

        dim_raw = _get(row, col_map, "dimension_class")
        dim_code = (
            _normalise_value("dimension_class", dim_raw, self._vn_idx)
            if dim_raw is not None
            else None
        )
        try:
            dim_enum = DimensionClass(dim_code) if dim_code else DimensionClass.UNDEFINED
        except ValueError:
            dim_enum = DimensionClass.UNDEFINED

        anomaly_id_raw = _get(row, col_map, "anomaly_id")
        anomaly_id = _to_str(anomaly_id_raw) or f"row{row_idx}"

        # Description: prefer the dedicated `description` column when the
        # vendor supplies one (NGP multi-sheet), else fall back to feature_type
        # text. This is what powers cluster-child detection.
        desc = (
            _to_str(_get(row, col_map, "description"))
            or _to_str(_get(row, col_map, "feature_type"))
            or ""
        )

        is_cluster_parent = fid_enum is FeatureIdentification.CORROSION_CLUSTER

        return Feature(
            anomaly_id=anomaly_id,
            source_run=run_id,
            source_row=row_idx,
            abs_distance_m=_to_float(_get(row, col_map, "abs_distance_m")) or 0.0,
            joint_number=joint_number,
            upstream_weld_dist_m=(
                uw_override
                if uw_override is not None
                else _to_float(_get(row, col_map, "upstream_weld_dist_m"))
            ),
            clock_decimal_hours=clock,
            latitude=_to_float(_get(row, col_map, "latitude")),
            longitude=_to_float(_get(row, col_map, "longitude")),
            altitude_m=_to_float(_get(row, col_map, "altitude_m")),
            wt_mm=wt,
            depth_pct_wt=depth_pct,
            length_mm=_to_float(_get(row, col_map, "length_mm")),
            width_mm=_to_float(_get(row, col_map, "width_mm")),
            surface=surface,
            feature_identification=fid_enum,
            dimension_class=dim_enum,
            raw_description=desc,
            vendor_erf=_to_float(_get(row, col_map, "erf")),
            vendor_psafe_kgcm2=_to_float(_get(row, col_map, "psafe")),
            is_cluster_parent=is_cluster_parent,
        )

    # ------------------------------------------------------------------

    def _build_joints(
        self,
        joints_data: dict[int, dict[str, Any]],
    ) -> list[Joint]:
        """Build Joint records from the full joint registry.

        For each unique joint_number we tracked while walking rows:
            abs_distance_start_m = min(start values seen)
            wt_mm                = mode of WT values across rows in this joint
            length_m             = first joint_length recorded (weld rows
                                   carry this in NGP single-sheet; Pipe-sheet
                                   rows carry it directly).
        """
        out: list[Joint] = []
        for jn, info in sorted(joints_data.items()):
            start = min(info["starts"]) if info["starts"] else 0.0
            wts = info["wts"]
            wt = Counter(wts).most_common(1)[0][0] if wts else None
            length = info["length"] or 0.0
            out.append(
                Joint(
                    joint_number=jn,
                    abs_distance_start_m=start,
                    length_m=length,
                    wt_mm=wt,
                )
            )
        return out

    # ------------------------------------------------------------------

    def _merge_joints_from_secondary_sheets(
        self,
        sheets: dict[str, list[list[Any]]],
        primary_sheet: str,
        joints_data: dict[int, dict[str, Any]],
    ) -> None:
        """Look for a 'Pipe' / 'Pipeline Tally' sheet (other than the primary
        defect sheet already walked) and merge any joint info from it.

        This is what makes the full joint count available downstream:
        NGP multi-sheet files put per-joint records on the `Pipe` sheet,
        and Athena multi-sheet puts welds/joints on `Pipeline Tally`.
        Defect sheets only ever list defect-bearing joints.
        """
        targets = ("pipe", "pipeline tally")
        for sn, rows in sheets.items():
            if sn == primary_sheet or not rows:
                continue
            sn_low = sn.casefold()
            if not any(t in sn_low for t in targets):
                continue
            idx, hits = _find_header_row(rows, self._syn_idx)
            if "joint_number" not in {self._syn_idx.get(_norm(c)) for c in rows[idx]}:
                continue
            col_map = _build_column_map(rows[idx], self._syn_idx)
            if "joint_number" not in col_map:
                continue
            ctx = None
            for row in rows[idx + 1 :]:
                jno = _to_int(_get(row, col_map, "joint_number"))
                if jno is not None:
                    ctx = jno
                if ctx is None:
                    continue
                entry = joints_data.setdefault(
                    ctx, {"starts": [], "wts": [], "length": None}
                )
                abs_d = _to_float(_get(row, col_map, "abs_distance_m"))
                if abs_d is not None:
                    uw = _to_float(_get(row, col_map, "upstream_weld_dist_m"))
                    entry["starts"].append(abs_d - uw if uw is not None else abs_d)
                wt = _to_float(_get(row, col_map, "wt_mm"))
                if wt is not None:
                    entry["wts"].append(wt)
                jlen = _to_float(_get(row, col_map, "joint_length_m"))
                if jlen is not None and entry["length"] is None:
                    entry["length"] = jlen
            # Walk only the first matching secondary sheet.
            return
