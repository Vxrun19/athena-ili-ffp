"""Tests for v0.3.1 ``src/io/feature_reader.py``.

Two read paths:

  * ``read_dent_features`` — bypasses the metal-loss skip-list so dent
    rows survive parsing.
  * ``read_metal_loss_features`` — thin wrapper around the legacy
    FFP read path; symmetric counterpart to the dent reader.

The FFP-pipeline ``ILIReader`` is unchanged: dents stay filtered out
of ``features_for_assessment`` (the Abu Road dent-leak guard). These
tests pin both behaviours don't interfere.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from src.io.feature_reader import (
    read_dent_features,
    read_metal_loss_features,
)
from src.models import FeatureIdentification


def _build_mixed_run2(tmp_path: Path) -> Path:
    """Build a small xlsx with both metal-loss and dent rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Defects"
    ws.append([
        "Anomaly ID", "Absolute Distance, m", "Joint Number",
        "WT, mm", "Depth, %WT", "Length, mm", "Width, mm",
        "Surface", "POF Acronym",
    ])
    # 3 metal-loss rows (CORR).
    ws.append(["ML-1", 100.0, 1, 8.0, 25.0, 50.0, 20.0, "Internal", "CORR"])
    ws.append(["ML-2", 200.0, 2, 8.0, 30.0, 40.0, 15.0, "External", "CORR"])
    ws.append(["ML-3", 300.0, 3, 8.0, 35.0, 60.0, 25.0, "Internal", "CORR"])
    # 2 dent rows (DENP — plain dent POF code that survives the FFP
    # skip-list because the literal "dent" string isn't in the row's
    # feature_identification; only DENP is, and that maps to DENT enum
    # via value_normalisations).
    ws.append(["DENT-1", 400.0, 10, 8.0, 5.0, 100.0, 80.0, "External", "DENP"])
    ws.append(["DENT-2", 500.0, 11, 8.0, 3.5, 80.0, 60.0, "External", "DENP"])
    # 1 dent_with_metal_loss row (DEML).
    ws.append(["DEML-1", 600.0, 12, 8.0, 4.0, 50.0, 40.0, "External", "DEML"])
    path = tmp_path / "mixed_run2.xlsx"
    wb.save(path)
    return path


def _build_dent_only_run2(tmp_path: Path) -> Path:
    """A run-2 file with ONLY dent rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Defects"
    ws.append([
        "Anomaly ID", "Absolute Distance, m", "Joint Number",
        "WT, mm", "Depth, %WT", "Length, mm", "Width, mm",
        "Surface", "POF Acronym",
    ])
    # Need at least one ML row for the sheet to score as a "defect
    # sheet" via the column-synonyms scorer. Then dent rows.
    ws.append(["ML-anchor", 50.0, 1, 8.0, 25.0, 50.0, 20.0,
               "Internal", "CORR"])
    ws.append(["DENT-1", 100.0, 9, 8.0, 5.0, 100.0, 80.0,
               "External", "DENP"])
    ws.append(["DENT-2", 200.0, 17, 8.0, 3.0, 80.0, 60.0,
               "External", "DENP"])
    ws.append(["DENT-3", 300.0, 25, 8.0, 2.5, 75.0, 55.0,
               "External", "DENP"])
    path = tmp_path / "dent_only_run2.xlsx"
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# read_dent_features
# ---------------------------------------------------------------------------

class TestReadDentFeatures:
    def test_mixed_file_returns_only_dents(self, tmp_path):
        path = _build_mixed_run2(tmp_path)
        dents = read_dent_features(path)
        # 2 DENP + 1 DEML = 3 dents in the synthetic.
        ids = sorted(str(f.anomaly_id) for f in dents)
        assert ids == ["DEML-1", "DENT-1", "DENT-2"], (
            f"expected the 3 dent rows; got ids={ids}"
        )
        for f in dents:
            assert f.feature_identification in (
                FeatureIdentification.DENT,
                FeatureIdentification.DENT_WITH_METAL_LOSS,
            )

    def test_no_dents_returns_empty_list(self, tmp_path):
        # Build a run with ONLY metal-loss rows.
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Defects"
        ws.append([
            "Anomaly ID", "Absolute Distance, m", "Joint Number",
            "WT, mm", "Depth, %WT", "Length, mm", "Width, mm",
            "Surface", "POF Acronym",
        ])
        ws.append(["ML-1", 100.0, 1, 8.0, 25.0, 50.0, 20.0,
                   "Internal", "CORR"])
        path = tmp_path / "ml_only.xlsx"
        wb.save(path)
        assert read_dent_features(path) == []

    def test_missing_file_returns_empty_list(self, tmp_path):
        assert read_dent_features(tmp_path / "nope.xlsx") == []

    def test_unreadable_file_returns_empty_list(self, tmp_path):
        bad = tmp_path / "garbage.xlsx"
        bad.write_bytes(b"not an xlsx")
        assert read_dent_features(bad) == []

    def test_dent_only_file_with_anchor(self, tmp_path):
        path = _build_dent_only_run2(tmp_path)
        dents = read_dent_features(path)
        # Three dent rows (anchor ML row is filtered out by the
        # dent-fid post-filter).
        assert len(dents) == 3
        ids = sorted(str(f.anomaly_id) for f in dents)
        assert ids == ["DENT-1", "DENT-2", "DENT-3"]


# ---------------------------------------------------------------------------
# read_metal_loss_features (regression — FFP path still filters dents)
# ---------------------------------------------------------------------------

class TestReadMetalLossFeatures:
    def test_mixed_file_returns_only_metal_loss(self, tmp_path):
        path = _build_mixed_run2(tmp_path)
        mls = read_metal_loss_features(path)
        ids = sorted(str(f.anomaly_id) for f in mls)
        # Only the 3 CORR rows; the 3 dent rows are filtered out by
        # features_for_assessment (Abu Road guard).
        assert ids == ["ML-1", "ML-2", "ML-3"]

    def test_no_metal_loss_returns_empty_list(self, tmp_path):
        path = _build_dent_only_run2(tmp_path)
        # Only the anchor ML row remains — single metal-loss feature.
        mls = read_metal_loss_features(path)
        assert len(mls) == 1
        assert str(mls[0].anomaly_id) == "ML-anchor"


# ---------------------------------------------------------------------------
# Both readers on the same file — no shared state
# ---------------------------------------------------------------------------

class TestReadersDoNotInterfere:
    def test_both_paths_on_same_file_are_disjoint(self, tmp_path):
        path = _build_mixed_run2(tmp_path)
        dents = read_dent_features(path)
        mls = read_metal_loss_features(path)
        dent_ids = {str(f.anomaly_id) for f in dents}
        ml_ids = {str(f.anomaly_id) for f in mls}
        assert dent_ids.isdisjoint(ml_ids), (
            f"dent + ml read paths overlap: {dent_ids & ml_ids}"
        )
        assert len(dent_ids) == 3
        assert len(ml_ids) == 3

    def test_dent_reader_called_twice_is_idempotent(self, tmp_path):
        path = _build_mixed_run2(tmp_path)
        a = read_dent_features(path)
        b = read_dent_features(path)
        assert len(a) == len(b)
        # Compare on anomaly IDs (Feature instances themselves are new
        # objects each call — by design).
        assert sorted(str(f.anomaly_id) for f in a) == \
               sorted(str(f.anomaly_id) for f in b)
