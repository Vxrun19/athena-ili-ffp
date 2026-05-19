"""Export-time tests for the Format Converter.

Prompt 34 surfaced a bug: `FormatConverter.convert()` re-read the source
file via `pd.read_excel()` at export time, which fails on `.csv` files
with the cryptic "Excel file format cannot be determined" error. The
load-time path (`_load_source`) handles CSVs correctly via the
encoding-fallback helper, but the export-time path went around it.

Fix Option A: cache the DataFrame at load time and pass it through to
`convert()` via the new `source_df=` / `pipe_df=` keyword arguments.

These tests pin:
  1. End-to-end: a latin-1 CSV with `°` in headers loads, the export
     produces a valid NGP-format xlsx — directly catches the original
     regression.
  2. The cache override actually bypasses disk I/O at export time
     (zero calls to read_csv / read_excel during convert).
  3. `read_source()` is now CSV-aware on its own, so callers that
     don't go through the GUI (CLI usage, direct API) get the same
     fix.
"""
from __future__ import annotations

from pathlib import Path
from unittest.mock import patch

import openpyxl
import pandas as pd
import pytest

from src.io.format_converter import FormatConverter, VendorProfile


def _write_latin1_csv(path: Path) -> Path:
    """Write a small Athena-2018-style CSV in latin-1 with `°` in headers.

    Headers deliberately AVOID embedded commas (a separate CSV-parsing
    edge case) so the test isolates the latin-1 encoding issue. The
    `°` character in "Latitude [°]" / "Longitude [°]" is 0xb0 — bare
    pd.read_csv() under utf-8 default chokes on it, which is the bug
    the encoding cascade fixes.
    """
    content = (
        "anomaly_id|abs_distance_m|joint_number|wt_mm|depth_pct_wt|"
        "length_mm|width_mm|surface|Latitude [°]|Longitude [°]\n"
        "1|100.0|5|7.1|25.0|50.0|20.0|Internal|23.15|70.20\n"
        "2|200.0|7|7.1|30.0|80.0|30.0|External|23.16|70.21\n"
        "3|300.0|9|7.1|15.0|40.0|15.0|Internal|23.17|70.22\n"
    )
    # latin-1: the ° symbol becomes 0xb0 — bare pd.read_csv chokes on it.
    with path.open("w", encoding="latin-1", newline="") as f:
        f.write(content)
    return path


# ---------------------------------------------------------------------------
# Test 1 — End-to-end: latin-1 CSV → NGP xlsx
# ---------------------------------------------------------------------------

class TestLatin1CsvExportEndToEnd:
    """The exact failure the user hit: load a latin-1 CSV with degree-symbol
    headers, export to NGP xlsx, no errors."""

    # Profile that maps every canonical field to the columns we wrote
    # in _write_latin1_csv (pipe-separated, so no comma-split ambiguity).
    _PROFILE_FOR_LATIN1_CSV = VendorProfile(
        vendor_name="Athena 2018 latin-1 CSV",
        sheet_name=None,                  # CSV has no sheets
        header_row=0,
        column_mappings={
            "anomaly_id":     "anomaly_id",
            "abs_distance_m": "abs_distance_m",
            "joint_number":   "joint_number",
            "wt_mm":          "wt_mm",
            "depth_pct_wt":   "depth_pct_wt",
            "length_mm":      "length_mm",
            "width_mm":       "width_mm",
            "surface":        "surface",
            "latitude":       "Latitude [°]",
            "longitude":      "Longitude [°]",
        },
        value_normalizations={
            "surface": {"Internal": "internal", "External": "external"},
        },
    )

    def test_convert_with_cached_source_df(self, tmp_path: Path):
        """The GUI's primary path: read the CSV via the encoding-fallback
        helper at load time, then pass the cached DataFrame into
        convert() via the new source_df= override."""
        csv = _write_latin1_csv(tmp_path / "athena_2018.csv")
        from src.io.format_converter.csv_input import (
            read_csv_with_encoding_fallback,
        )
        # Read with pipe separator so the "," inside "Latitude [°]"-
        # adjacent headers doesn't matter. header=0 means row 0
        # becomes column names.
        body = read_csv_with_encoding_fallback(csv, header=0, sep="|")
        assert "Latitude [°]" in body.columns, (
            f"degree-symbol header lost from CSV decode: {list(body.columns)}"
        )

        out_path = tmp_path / "out.xlsx"
        # Pass the cached DataFrame so convert() doesn't re-read.
        result = FormatConverter(self._PROFILE_FOR_LATIN1_CSV).convert(
            csv, out_path, source_df=body,
        )
        assert result == out_path
        assert out_path.exists() and out_path.stat().st_size > 0

        # Verify the output is a real xlsx with a "Defects" sheet and
        # the canonical NGP column names (not the source's snake_case).
        wb = openpyxl.load_workbook(out_path, read_only=True)
        try:
            assert "Defects" in wb.sheetnames
            ws = wb["Defects"]
            header_cells = [c.value for c in next(ws.iter_rows())]
            # NGP canonical columns from converter.NGP_OUTPUT_COLUMNS.
            assert "Absolute Distance, m" in header_cells
            assert "Surface" in header_cells
        finally:
            wb.close()

    def test_convert_without_cache_reads_csv_directly(self, tmp_path: Path):
        """Defense-in-depth: even without the GUI's caching, convert()
        on a .csv path should succeed because read_source() is now
        polymorphic by extension (the v0.2 code's bare pd.read_excel
        would raise 'Excel file format cannot be determined' here)."""
        csv = _write_latin1_csv(tmp_path / "athena_2018.csv")
        # read_source uses pandas default sep="," — so we need a CSV
        # with comma separators for this path. Rewrite the file using
        # commas (no embedded commas in headers this time).
        csv.write_text(
            "anomaly_id,abs_distance_m,joint_number,wt_mm,depth_pct_wt,"
            "length_mm,width_mm,surface,Latitude [°],Longitude [°]\n"
            "1,100.0,5,7.1,25.0,50.0,20.0,Internal,23.15,70.20\n"
            "2,200.0,7,7.1,30.0,80.0,30.0,External,23.16,70.21\n",
            encoding="latin-1",
        )
        out_path = tmp_path / "no_cache_out.xlsx"
        # No source_df= → convert() routes through read_source(),
        # which must now dispatch to the CSV helper.
        result = FormatConverter(self._PROFILE_FOR_LATIN1_CSV).convert(
            csv, out_path,
        )
        assert result.exists()
        wb = openpyxl.load_workbook(out_path, read_only=True)
        try:
            assert "Defects" in wb.sheetnames
        finally:
            wb.close()


# ---------------------------------------------------------------------------
# Test 2 — Cache is honoured: zero disk reads during convert when DFs cached
# ---------------------------------------------------------------------------

class TestCacheReuse:
    """When the GUI passes source_df=, convert() must NOT touch the disk."""

    def test_source_df_override_skips_read_source(self, tmp_path: Path):
        # A bogus source path that would fail if convert() tried to read it.
        bogus_path = tmp_path / "does_not_exist.csv"
        # But we pass a cached DataFrame, so read_source() should be
        # skipped entirely.
        cached = pd.DataFrame({
            "anomaly_id":     ["1", "2"],
            "abs_distance_m": [100.0, 200.0],
            "joint_number":   [5, 7],
            "wt_mm":          [7.1, 7.1],
            "depth_pct_wt":   [25.0, 30.0],
            "length_mm":      [50.0, 80.0],
            "width_mm":       [20.0, 30.0],
            "surface":        ["internal", "external"],
        })
        profile = VendorProfile(
            vendor_name="cache-test",
            sheet_name="Defects",
            header_row=0,
            column_mappings={
                "anomaly_id":      "anomaly_id",
                "abs_distance_m":  "abs_distance_m",
                "joint_number":    "joint_number",
                "wt_mm":           "wt_mm",
                "depth_pct_wt":    "depth_pct_wt",
                "length_mm":       "length_mm",
                "width_mm":        "width_mm",
                "surface":         "surface",
            },
        )
        out_path = tmp_path / "cached_out.xlsx"
        # Patch the I/O helpers to assert they're NEVER called.
        with patch("src.io.format_converter.converter.pd.read_excel") as mock_xl, \
             patch(
                 "src.io.format_converter.csv_input."
                 "read_csv_with_encoding_fallback"
             ) as mock_csv:
            FormatConverter(profile).convert(
                bogus_path, out_path, source_df=cached,
            )
            mock_xl.assert_not_called()
            mock_csv.assert_not_called()
        assert out_path.exists()

    def test_pipe_df_override_skips_read_pipe_source(self, tmp_path: Path):
        """Same for the pipe-sheet pass-through."""
        bogus_path = tmp_path / "does_not_exist.xlsx"
        cached_defects = pd.DataFrame({
            "anomaly_id":     ["1"],
            "abs_distance_m": [100.0],
            "joint_number":   [5],
            "wt_mm":          [7.1],
            "depth_pct_wt":   [25.0],
            "length_mm":      [50.0],
            "width_mm":       [20.0],
            "surface":        ["internal"],
        })
        cached_pipe = pd.DataFrame({
            "joint_number":   [5, 6, 7],
            "joint_length_m": [12.0, 12.0, 12.0],
        })
        profile = VendorProfile(
            vendor_name="pipe-cache-test",
            sheet_name="Defects",
            header_row=0,
            column_mappings={
                "anomaly_id":      "anomaly_id",
                "abs_distance_m":  "abs_distance_m",
                "joint_number":    "joint_number",
                "wt_mm":           "wt_mm",
                "depth_pct_wt":    "depth_pct_wt",
                "length_mm":       "length_mm",
                "width_mm":        "width_mm",
                "surface":         "surface",
            },
            pipe_sheet_name="Pipeline Tally",
            pipe_header_row=0,
            pipe_column_mappings={
                "joint_number":   "joint_number",
                "joint_length_m": "joint_length_m",
            },
        )
        out_path = tmp_path / "pipe_cached_out.xlsx"
        with patch("src.io.format_converter.converter.pd.read_excel") as mock_xl:
            FormatConverter(profile).convert(
                bogus_path, out_path,
                source_df=cached_defects,
                pipe_df=cached_pipe,
            )
            mock_xl.assert_not_called()
        assert out_path.exists()
        # The output xlsx should carry BOTH sheets.
        wb = openpyxl.load_workbook(out_path, read_only=True)
        try:
            assert "Defects" in wb.sheetnames
            assert "Pipeline Tally" in wb.sheetnames
        finally:
            wb.close()


# ---------------------------------------------------------------------------
# Test 3 — Re-load invalidates the cache
# ---------------------------------------------------------------------------

class TestReloadInvalidatesCache:
    """When the user loads a different file via Browse, the GUI's cached
    DataFrame must update — exporting after the second load uses the
    second file's data, NOT the first.

    This isn't really about FormatConverter (which is stateless wrt the
    cache); it's about the GUI's `_source_df` rebinding. We drive
    the GUI off-screen so the regression coverage is realistic.
    """

    def test_second_load_replaces_source_df(self, tmp_path: Path):
        import os
        os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
        from PyQt6.QtWidgets import QApplication
        from src.gui.screens.format_converter import FormatConverterScreen

        app = QApplication.instance() or QApplication([])

        # File A — minimal Athena/NGP-style xlsx (use xlsx not csv so
        # both files are normal Excel; we're testing the cache, not
        # the CSV path here).
        from openpyxl import Workbook
        path_a = tmp_path / "first.xlsx"
        path_b = tmp_path / "second.xlsx"
        for p, marker_value in ((path_a, 1.0), (path_b, 2.0)):
            wb = Workbook()
            ws = wb.active
            ws.title = "Defects"
            ws.append(["anomaly_id", "abs_distance_m", "joint_number",
                       "wt_mm", "depth_pct_wt", "length_mm", "width_mm",
                       "surface"])
            ws.append(["1", marker_value, 1, 7.1, 25.0, 50.0, 20.0,
                       "internal"])
            wb.save(p)

        screen = FormatConverterScreen()
        screen.show()
        try:
            screen._load_source(path_a)
            QApplication.processEvents()
            df_a = screen._source_df.copy()
            assert float(df_a.iloc[0]["abs_distance_m"]) == 1.0

            screen._load_source(path_b)
            QApplication.processEvents()
            df_b = screen._source_df.copy()
            assert float(df_b.iloc[0]["abs_distance_m"]) == 2.0

            # Confirm the cache really did rebind — same instance check
            # would be fragile (pandas may return identical objects in
            # some corner cases); compare contents instead.
            assert not df_a.equals(df_b)
        finally:
            screen.close()
