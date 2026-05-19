"""Tests for src.reports.annexure_writer.

We generate the Excel output and then inspect it directly with openpyxl
to validate sheets, merges, header content, row counts, sort order, and
specific data values for the canonical Kandla #125 case.

The reference files in `/examples/` (FFP_Report_IPS_1_to_IPS2_Annexure_E__F__r4.xlsx
and IPS__Samakhiali...xls) drove the layout decisions documented in
`docs/REPORT_FORMATS.md`.
"""
from __future__ import annotations

from datetime import date
from pathlib import Path

import openpyxl
import pytest

from src.core.cgr import CGRCalculator
from src.core.defect_matcher import DefectMatcher
from src.core.ffp import ffp_assess
from src.core.joint_alignment import JointAligner
from src.core.repair_predictor import RepairPredictor
from src.io.ili_reader import ILIReader
from src.models import MAOPZone, Pipeline, Project
from src.reports.annexure_writer import AnnexureWriter
from src.validation.flag_aggregator import FlagAggregator

EXAMPLES = Path(__file__).resolve().parents[1] / "examples"


# ---------------------------------------------------------------------------
# Fixture: full Kandla pipeline + outputs
# ---------------------------------------------------------------------------

@pytest.fixture(scope="module")
def kandla_outputs():
    """Run the full chain on the Kandla pair, return everything the
    AnnexureWriter consumes."""
    pipeline = Pipeline(
        pipeline_name='Kandla-Samakhiali 10" LPG',
        client_name="GAIL (India) Limited",
        diameter_mm=273.0, smys_mpa=358.0,
        maop_zones=[MAOPZone(6.0, 8.0, 0.72, 70.0)],
    )
    reader = ILIReader()
    run1 = reader.read(
        EXAMPLES / "10_inch_Kandla_to_Samakhiali__58_2_km_FR_Pipe_Tally_Rev_0.xlsx",
        run_id="k1",
    )
    run1.inspection_date = date(2018, 12, 15)
    run2 = reader.read(EXAMPLES / "1ZSV_Pipeline_Listing.xlsx", run_id="k2")
    run2.inspection_date = date(2023, 3, 15)
    project = Project(
        project_name="FFP_Kandla_Samakhiali",
        pipeline=pipeline, run_1=run1, run_2=run2,
    )
    ja = JointAligner().align(run1, run2)
    mr = DefectMatcher().match(run1, run2, ja.matches)
    cgrs = CGRCalculator({"mode": "hybrid"}).compute(mr, years_between=4.25)
    ffps_by_id = {}
    for c in cgrs:
        fl = ffp_assess(c.feature, pipeline)
        ffps_by_id[c.feature.anomaly_id] = next(
            (f for f in fl if f.is_controlling), fl[0]
        )
    preds = RepairPredictor().predict(
        cgrs, ffps_by_id, pipeline,
        run2_inspection_date=date(2023, 3, 15),
    )
    flag_report = FlagAggregator().aggregate(
        run1=run1, run2=run2, joint_alignment=ja, match_result=mr,
        cgr_results=cgrs, ffp_results=list(ffps_by_id.values()),
        predictions=preds,
    )
    return {
        "project": project, "pipeline": pipeline,
        "cgr_results": cgrs,
        "ffp_results": list(ffps_by_id.values()),
        "predictions": preds,
        "flag_report": flag_report,
    }


@pytest.fixture
def kandla_ef_path(tmp_path, kandla_outputs):
    """Write the E/F annexure and return the path. New file per test."""
    out = tmp_path / "kandla_ef.xlsx"
    AnnexureWriter().write(
        cgr_results=kandla_outputs["cgr_results"],
        ffp_results=kandla_outputs["ffp_results"],
        repair_predictions=kandla_outputs["predictions"],
        flag_report=kandla_outputs["flag_report"],
        project=kandla_outputs["project"],
        pipeline=kandla_outputs["pipeline"],
        output_path=str(out),
        format="E_F",
    )
    return out


# ---------------------------------------------------------------------------
# Format E/F — Annexure E
# ---------------------------------------------------------------------------

class TestKandlaAnnexureE:
    def test_file_created_and_sheets_present(self, kandla_ef_path):
        wb = openpyxl.load_workbook(kandla_ef_path)
        assert "Annexure E" in wb.sheetnames
        assert "Annexure F" in wb.sheetnames
        assert "QA Issues" in wb.sheetnames

    def test_title_row_text_and_merge(self, kandla_ef_path):
        wb = openpyxl.load_workbook(kandla_ef_path)
        ws = wb["Annexure E"]
        # Title text matches the reference convention.
        assert ws["A1"].value == "Annexure E: Run to Run Comparison"
        # v0.2.6: layout widened from 13 to 14 columns (added
        # "CGR raw (mm/yr)" before "CGR (mm/yr)"). Title merge is
        # now A1:N1.
        merges = {r.coord for r in ws.merged_cells.ranges}
        assert "A1:N1" in merges

    def test_group_header_row_merges(self, kandla_ef_path):
        wb = openpyxl.load_workbook(kandla_ef_path)
        ws = wb["Annexure E"]
        merges = {r.coord for r in ws.merged_cells.ranges}
        # Exact 5 group-header merges per the reference file.
        for expected in ("A2:D2", "E2:F2", "G2:H2", "I2:J2", "K2:L2"):
            assert expected in merges, f"missing merge {expected}; got {merges}"

    def test_group_header_text(self, kandla_ef_path):
        wb = openpyxl.load_workbook(kandla_ef_path)
        ws = wb["Annexure E"]
        # Year labels from the project's inspection_date.
        assert ws["A2"].value == "Feature Detail as per ILI 2023"
        assert ws["E2"].value == "Abs. Distance, (m)"
        assert ws["G2"].value == "Anomaly Depth, (%)"
        assert ws["I2"].value == "Anomaly Orientation"
        assert ws["K2"].value == "Anomaly Location"

    def test_field_header_row(self, kandla_ef_path):
        wb = openpyxl.load_workbook(kandla_ef_path)
        ws = wb["Annexure E"]
        # Row 3 — field headers. v0.2.6: 14 columns; the new
        # "CGR raw (mm/yr)" sits immediately before "CGR (mm/yr)".
        expected = [
            "S.N.", "Anomaly ID", "Wall Thickness, (mm)", "Joint Number",
            "ILI 2023", "ILI 2018",
            "ILI 2023", "ILI 2018",
            "ILI 2023", "ILI 2018",
            "ILI 2023", "ILI 2018",
            "CGR raw (mm/yr)",
            "CGR (mm/yr)",
        ]
        actual = [ws.cell(3, c).value for c in range(1, 15)]
        assert actual == expected

    def test_row_count_matches_total_run2_features(self, kandla_ef_path, kandla_outputs):
        wb = openpyxl.load_workbook(kandla_ef_path)
        ws = wb["Annexure E"]
        # 3 header rows + 333 features (22 matched + 311 unmatched) = 336.
        n_data_rows = ws.max_row - 3
        assert n_data_rows == len(kandla_outputs["cgr_results"]) == 333

    def test_data_rows_sorted_by_abs_distance(self, kandla_ef_path):
        wb = openpyxl.load_workbook(kandla_ef_path)
        ws = wb["Annexure E"]
        # Column E is run-2 abs_distance.
        distances = [
            float(ws.cell(r, 5).value)
            for r in range(4, ws.max_row + 1)
            if ws.cell(r, 5).value is not None
        ]
        assert distances == sorted(distances), \
            "Annexure E data not sorted by run-2 abs_distance ascending"

    def test_feature_125_row_values(self, kandla_ef_path):
        """The canonical highest-CGR defect — must match the validated
        numbers from the full chain (CGR=0.2522, depth pair 28.75/12,
        orientation 05:08/05:18, surfaces int./int.)."""
        wb = openpyxl.load_workbook(kandla_ef_path)
        ws = wb["Annexure E"]
        target_row = None
        for r in range(4, ws.max_row + 1):
            if ws.cell(r, 2).value == "125":
                target_row = r
                break
        assert target_row is not None, "feature #125 not in Annexure E"
        # WT
        assert ws.cell(target_row, 3).value == pytest.approx(6.4)
        # Joint number
        assert ws.cell(target_row, 4).value == 6410
        # Abs distance run-2 / run-1
        assert ws.cell(target_row, 5).value == pytest.approx(7453.053, abs=0.005)
        assert ws.cell(target_row, 6).value == pytest.approx(7426.979, abs=0.005)
        # Depth pct run-2 / run-1
        assert ws.cell(target_row, 7).value == pytest.approx(28.75, abs=0.01)
        assert ws.cell(target_row, 8).value == pytest.approx(12.0, abs=0.5)
        # Orientation as hh:mm:ss text
        assert ws.cell(target_row, 9).value == "05:08:00"
        assert ws.cell(target_row, 10).value == "05:18:00"
        # Surface
        assert ws.cell(target_row, 11).value == "int."
        assert ws.cell(target_row, 12).value == "int."
        # CGR
        assert ws.cell(target_row, 13).value == pytest.approx(0.2522, abs=0.0002)

    def test_number_formats_applied(self, kandla_ef_path):
        wb = openpyxl.load_workbook(kandla_ef_path)
        ws = wb["Annexure E"]
        # First data row — column-specific number formats
        assert ws.cell(4, 3).number_format == "0.0"      # WT
        assert ws.cell(4, 5).number_format == "0.000"    # abs distance
        assert ws.cell(4, 7).number_format == "0.00"     # depth pct
        assert ws.cell(4, 13).number_format == "0.0000"  # CGR


# ---------------------------------------------------------------------------
# Format E/F — Annexure F
# ---------------------------------------------------------------------------

class TestKandlaAnnexureF:
    def test_three_title_rows(self, kandla_ef_path):
        wb = openpyxl.load_workbook(kandla_ef_path)
        ws = wb["Annexure F"]
        assert ws["A1"].value == "Annexure F: Metal Loss Anomalies"
        # Pipeline name on row 2
        assert "Kandla" in (ws["A2"].value or "") \
            or "GAIL" in (ws["A2"].value or "")
        # Section/title on row 3 (may also have content)
        assert ws["A3"].value is not None

    def test_field_header_row(self, kandla_ef_path):
        wb = openpyxl.load_workbook(kandla_ef_path)
        ws = wb["Annexure F"]
        headers = [ws.cell(4, c).value for c in range(1, 17)]
        # Spot-check key columns
        assert headers[0] == "S.N."
        assert headers[1] == "Feature ID"
        assert "Absolute Distance" in (headers[2] or "")
        assert headers[3] == "Latitude"
        assert headers[4] == "Longitude"
        assert headers[5] == "Joint No."
        assert headers[8] == "Event"
        assert headers[9] == "Surface"
        assert "Orientation" in (headers[11] or "")
        assert "Repair" in (headers[15] or "")

    def test_data_rows_present_and_sorted(self, kandla_ef_path, kandla_outputs):
        wb = openpyxl.load_workbook(kandla_ef_path)
        ws = wb["Annexure F"]
        n_data_rows = ws.max_row - 4
        assert n_data_rows == len(kandla_outputs["cgr_results"])
        # Abs distance column 3 (per layout)
        distances = [
            float(ws.cell(r, 3).value)
            for r in range(5, ws.max_row + 1)
            if ws.cell(r, 3).value is not None
        ]
        assert distances == sorted(distances)

    def test_feature_125_row_in_annexure_f(self, kandla_ef_path):
        wb = openpyxl.load_workbook(kandla_ef_path)
        ws = wb["Annexure F"]
        target_row = None
        for r in range(5, ws.max_row + 1):
            if ws.cell(r, 2).value == "125":
                target_row = r
                break
        assert target_row is not None
        assert ws.cell(target_row, 3).value == pytest.approx(7453.053, abs=0.005)
        assert ws.cell(target_row, 6).value == 6410
        assert ws.cell(target_row, 10).value == "int."
        assert ws.cell(target_row, 11).value == pytest.approx(6.4)
        assert ws.cell(target_row, 12).value == "05:08:00"
        assert ws.cell(target_row, 13).value == pytest.approx(28.75, abs=0.01)
        # Repair date — Kandla predicts NONE_WITHIN_HORIZON → "After ..."
        repair = ws.cell(target_row, 16).value
        assert isinstance(repair, str) and "After" in repair


# ---------------------------------------------------------------------------
# QA Issues sheet
# ---------------------------------------------------------------------------

class TestQAIssuesSheet:
    def test_summary_text_present(self, kandla_ef_path):
        wb = openpyxl.load_workbook(kandla_ef_path)
        ws = wb["QA Issues"]
        assert ws["A1"].value == "QA Issues"
        # Row 2 has the summary string from FlagReport.summary
        assert "finding" in (ws["A2"].value or "").lower() \
            or "clean" in (ws["A2"].value or "").lower()

    def test_flag_rows_listed(self, kandla_ef_path, kandla_outputs):
        wb = openpyxl.load_workbook(kandla_ef_path)
        ws = wb["QA Issues"]
        # Row 3 is the column-header row; rows 4+ are data.
        n_flag_rows = max(0, ws.max_row - 3)
        assert n_flag_rows >= 1
        # The aggregator emits ≥311 UNMATCHED_RUN2 flags on Kandla, plus
        # NO_CLUSTERS_IN_EITHER_RUN, LOW_DEFECT_MATCH_RATE, etc.
        # Just verify the table has flag rows.

    def test_severity_ordering_error_then_warn_then_info(self, kandla_ef_path):
        wb = openpyxl.load_workbook(kandla_ef_path)
        ws = wb["QA Issues"]
        severities = [
            ws.cell(r, 1).value or ""
            for r in range(4, ws.max_row + 1)
        ]
        order = {"ERROR": 0, "WARN": 1, "INFO": 2}
        rank = [order.get(s, 9) for s in severities]
        assert rank == sorted(rank), "QA Issues not sorted ERROR -> WARN -> INFO"


# ---------------------------------------------------------------------------
# Format B/C/D — older GAIL format
# ---------------------------------------------------------------------------

class TestKandlaAnnexureBCD:
    """Use the same Kandla data but write in the B/C/D format (sanity
    check that the older layout also generates without errors and has
    the right sheets/structure)."""

    @pytest.fixture
    def bcd_path(self, tmp_path, kandla_outputs):
        out = tmp_path / "kandla_bcd.xlsx"
        AnnexureWriter().write(
            cgr_results=kandla_outputs["cgr_results"],
            ffp_results=kandla_outputs["ffp_results"],
            repair_predictions=kandla_outputs["predictions"],
            project=kandla_outputs["project"],
            pipeline=kandla_outputs["pipeline"],
            output_path=str(out),
            format="B_C_D",
        )
        return out

    def test_sheets_present(self, bcd_path):
        wb = openpyxl.load_workbook(bcd_path)
        assert "Annexure B" in wb.sheetnames
        assert "Annexure C" in wb.sheetnames
        assert "Annexure D" in wb.sheetnames

    def test_annexure_b_only_contains_matched(self, bcd_path, kandla_outputs):
        wb = openpyxl.load_workbook(bcd_path)
        ws = wb["Annexure B"]
        # Annexure B is matched-only. Kandla has 22 matched features.
        n_matched = sum(
            1 for c in kandla_outputs["cgr_results"]
            if c.matched_to_run1 is not None
        )
        # Headers occupy rows 1-7; data starts row 8.
        n_data = max(0, ws.max_row - 7)
        assert n_data == n_matched

    def test_annexure_c_has_b31g_in_subtitle(self, bcd_path):
        wb = openpyxl.load_workbook(bcd_path)
        ws = wb["Annexure C"]
        assert "B-31G" in (ws["A2"].value or "")

    def test_annexure_d_has_kastner_in_subtitle(self, bcd_path):
        wb = openpyxl.load_workbook(bcd_path)
        ws = wb["Annexure D"]
        assert "Kastner" in (ws["A2"].value or "")

    def test_feature_125_in_annexure_b_with_cgr(self, bcd_path):
        wb = openpyxl.load_workbook(bcd_path)
        ws = wb["Annexure B"]
        target_row = None
        for r in range(8, ws.max_row + 1):
            if ws.cell(r, 2).value == "125":
                target_row = r
                break
        assert target_row is not None
        # CGR is the last column (column 10).
        assert ws.cell(target_row, 10).value == pytest.approx(0.2522, abs=0.0002)


# ---------------------------------------------------------------------------
# Misc — error handling, etc.
# ---------------------------------------------------------------------------

class TestWriterErrors:
    def test_unknown_format_raises(self, tmp_path, kandla_outputs):
        out = tmp_path / "bogus.xlsx"
        with pytest.raises(ValueError, match="unknown format"):
            AnnexureWriter().write(
                cgr_results=kandla_outputs["cgr_results"],
                ffp_results=kandla_outputs["ffp_results"],
                repair_predictions=kandla_outputs["predictions"],
                project=kandla_outputs["project"],
                output_path=str(out),
                format="Z_X",
            )

    def test_empty_inputs_generates_just_headers(self, tmp_path):
        out = tmp_path / "empty.xlsx"
        AnnexureWriter().write(
            cgr_results=[], ffp_results=[], repair_predictions=[],
            output_path=str(out), format="E_F",
        )
        wb = openpyxl.load_workbook(out)
        ws = wb["Annexure E"]
        # 3 header rows, 0 data rows
        assert ws.max_row == 3
