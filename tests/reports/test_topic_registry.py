"""Tests for src/reports/topic_registry.py (v0.2.5).

The topic registry defines which annexure-sheets are available to the
report builder. These tests pin the spec'd 7-topic set and its core
invariants.
"""
from __future__ import annotations

import pytest

from src.reports.topic_registry import (
    AnnexureTopic,
    TOPIC_REGISTRY,
    all_topics_in_order,
    default_annexure_selection,
)


# ---------------------------------------------------------------------------
# 7-topic inventory
# ---------------------------------------------------------------------------

# Spec from v0.2.5 task description.
_EXPECTED_TOPICS: list[tuple[str, str, bool]] = [
    # (id, default_letter, implemented)
    ("guidelines_formulas",     "A", True),
    ("results_ili_comparison",  "B", True),
    ("metal_loss_anomalies",    "C", True),
    ("estimated_erf_defects",   "D", True),
    ("estimated_erf_circ",      "E", True),
    # v0.3.1: dent strain now implemented (was placeholder in v0.2.5).
    ("dent_strain_b318",        "F", True),
    ("qa_findings",             "G", True),
]


class TestRegistryStructure:
    def test_registry_contains_exactly_seven_topics(self):
        assert len(TOPIC_REGISTRY) == 7

    def test_all_topics_are_AnnexureTopic_instances(self):
        for topic in TOPIC_REGISTRY.values():
            assert isinstance(topic, AnnexureTopic)

    def test_topic_dataclass_is_frozen(self):
        topic = next(iter(TOPIC_REGISTRY.values()))
        with pytest.raises(Exception):
            # frozen=True -> attribute assignment fails
            topic.display_name = "mutated"     # type: ignore[misc]

    def test_topics_have_unique_ids(self):
        ids = [t.id for t in all_topics_in_order()]
        assert len(set(ids)) == len(ids), (
            f"Topic IDs are not unique: {ids}"
        )

    def test_topics_have_unique_default_letters(self):
        letters = [t.default_letter for t in all_topics_in_order()]
        assert len(set(letters)) == len(letters), (
            f"Default letters are not unique: {letters}"
        )

    @pytest.mark.parametrize("tid,letter,implemented", _EXPECTED_TOPICS)
    def test_each_spec_topic_present(self, tid, letter, implemented):
        assert tid in TOPIC_REGISTRY, f"missing topic {tid!r}"
        t = TOPIC_REGISTRY[tid]
        assert t.default_letter == letter, (
            f"{tid}: default_letter mismatch (got {t.default_letter!r}, "
            f"expected {letter!r})"
        )
        assert t.implemented == implemented, (
            f"{tid}: implemented mismatch (got {t.implemented}, "
            f"expected {implemented})"
        )

    def test_all_topics_implemented_in_v031(self):
        """v0.3.1: dent strain joined the implemented list. All seven
        topics now produce real output (no placeholders)."""
        unimpl = [
            t.id for t in TOPIC_REGISTRY.values() if not t.implemented
        ]
        assert unimpl == [], (
            f"v0.3.1 expects every topic implemented; got unimpl={unimpl}"
        )

    def test_writers_are_callable(self):
        for topic in TOPIC_REGISTRY.values():
            assert callable(topic.writer), (
                f"{topic.id}: writer is not callable"
            )

    def test_display_order_matches_default_letter_alphabetical(self):
        """The canonical display order is also A → G in default letters."""
        topics = all_topics_in_order()
        letters = [t.default_letter for t in topics]
        assert letters == sorted(letters), (
            f"Display order does not match alphabetical default letters: {letters}"
        )

    def test_all_topics_in_order_returns_fresh_list(self):
        """Mutating the returned list must not mutate the registry."""
        first = all_topics_in_order()
        first.clear()
        second = all_topics_in_order()
        assert len(second) == 7


# ---------------------------------------------------------------------------
# Default selection (legacy E_F preset equivalent)
# ---------------------------------------------------------------------------

class TestDefaultAnnexureSelection:
    def test_returns_three_topics(self):
        sel = default_annexure_selection()
        assert len(sel) == 3

    def test_legacy_topic_ids(self):
        sel = default_annexure_selection()
        ids = [tid for tid, _letter in sel]
        assert ids == [
            "results_ili_comparison",
            "metal_loss_anomalies",
            "qa_findings",
        ]

    def test_uses_default_letters(self):
        sel = default_annexure_selection()
        # Letters come from each topic's default_letter.
        for tid, letter in sel:
            assert TOPIC_REGISTRY[tid].default_letter == letter

    def test_returns_fresh_list_per_call(self):
        a = default_annexure_selection()
        b = default_annexure_selection()
        assert a is not b
        a.clear()
        assert len(b) == 3


# ---------------------------------------------------------------------------
# v0.3.3: estimated_erf_circ writer — Kastner row emission for every
# eligible circumferential feature (not just where Kastner controls)
# ---------------------------------------------------------------------------

class TestEstimatedErfCircWriter:
    """Pins the v0.3.3 fix: a mixed feature population yields one Kastner
    row per circumferentially-classified feature, regardless of whether
    Kastner's Psafe is the lower (controlling) value vs the primary
    method. Pre-v0.3.3 this sheet was empty for BPCL Mathura-Piyala
    (323 circ defects → 0 rows) because the filter required ``method is
    FFPMethod.KASTNER`` and ``ffps_by_id`` only stored controlling
    results.
    """

    def _build_results(self):
        """Build a synthetic AnalysisResult-shaped namespace with a
        mixed feature population: 5 POF-circ-labeled, 3 geometric-circ
        (POF UNDEFINED, W > L), 7 axial, 10 pitting.

        Returns the SimpleNamespace + a list of expected Kastner-eligible
        anomaly_ids.
        """
        from datetime import date
        from types import SimpleNamespace

        from src.models import (
            DimensionClass, Feature, FeatureIdentification, Surface,
        )
        from src.core.cgr import CGRResult

        def mk(anomaly_id, dim, length, width, raw_desc=""):
            f = Feature(
                anomaly_id=str(anomaly_id),
                source_run="run_2",
                source_row=int(anomaly_id),
                abs_distance_m=float(anomaly_id),
                joint_number=10,
                dimension_class=dim,
                raw_description=raw_desc,
                length_mm=length, width_mm=width,
                wt_mm=6.4, depth_pct_wt=15.0,
                feature_identification=FeatureIdentification.CORROSION,
                surface=Surface.EXTERNAL,
            )
            return CGRResult(
                feature=f, matched_to_run1=None,
                cgr_mm_yr=0.05, feature_cgr_mm_yr=0.05,
                mode_used="population_floor",
                depth_old_used_mm=10.0 * 6.4 / 100.0,
                depth_new_mm=15.0 * 6.4 / 100.0,
                years_between=5.0,
            )

        # 5 POF-circ-labeled (CISL/CIGR)
        circ_pof = [
            mk(101, DimensionClass.CIRCUMFERENTIAL_SLOTTING, 8.0, 20.0),
            mk(102, DimensionClass.CIRCUMFERENTIAL_SLOTTING, 10.0, 15.0),
            mk(103, DimensionClass.CIRCUMFERENTIAL_GROOVING, 12.0, 30.0),
            mk(104, DimensionClass.CIRCUMFERENTIAL_GROOVING, 9.0, 18.0),
            mk(105, DimensionClass.CIRCUMFERENTIAL_SLOTTING, 14.0, 25.0),
        ]
        # 3 geometric-circ (POF UNDEFINED but W > L)
        geom_circ = [
            mk(201, DimensionClass.UNDEFINED, 10.0, 30.0),
            mk(202, DimensionClass.UNDEFINED, 12.0, 40.0),
            mk(203, DimensionClass.UNDEFINED, 8.0, 22.0),
        ]
        # 7 axial — must NOT be picked.
        axial = [
            mk(301 + i, DimensionClass.AXIAL_SLOTTING, 30.0, 10.0)
            for i in range(7)
        ]
        # 10 pitting — must NOT be picked even if some have W > L
        # (POF override is what we're testing here).
        pitting = [
            mk(401 + i, DimensionClass.PITTING,
               length=20.0 if i % 2 == 0 else 5.0,
               width=10.0  if i % 2 == 0 else 25.0)   # half have W > L
            for i in range(10)
        ]
        cgrs = circ_pof + geom_circ + axial + pitting

        results = SimpleNamespace(
            cgr_results=cgrs,
            ffp_results=[],
            repair_predictions=[],
            flag_report=None,
        )
        expected_eligible_ids = {
            "101", "102", "103", "104", "105",
            "201", "202", "203",
        }
        return results, expected_eligible_ids

    def _build_project(self):
        from src.models import MAOPZone, Pipeline, Project
        pipeline = Pipeline(
            pipeline_name="test", diameter_mm=406.0, length_km=10.0,
            install_year=2010, material_grade="API 5L X60", smys_mpa=413.0,
            maop_zones=[
                MAOPZone(
                    wt_mm_min=5.9, wt_mm_max=7.6,
                    design_factor=0.72, maop_kgcm2=88.0,
                ),
            ],
        )
        return Project(
            project_name="kastner-mixed-test",
            pipeline=pipeline,
        )

    def test_eight_features_emitted_with_mixed_population(self):
        """5 POF-circ + 3 geom-circ + 7 axial + 10 pitting → 8 rows."""
        from openpyxl import Workbook
        from src.reports.topic_registry import _topic_estimated_erf_circ

        results, expected_ids = self._build_results()
        project = self._build_project()
        wb = Workbook()
        wb.remove(wb.active)

        _topic_estimated_erf_circ(
            wb, "Annexure E", project, results, run2_year=2023,
            title_text="Annexure E — Kastner",
        )

        ws = wb["Annexure E"]
        # Header is 4 rows; data starts at row 5. Count data rows by
        # looking for non-empty S.N. column.
        data_rows = 0
        emitted_ids: list[str] = []
        for row_idx in range(5, ws.max_row + 1):
            sn = ws.cell(row_idx, 1).value
            fid = ws.cell(row_idx, 2).value
            if sn is None:
                continue
            data_rows += 1
            if fid is not None:
                emitted_ids.append(str(fid))

        assert data_rows == 8, (
            f"expected 8 Kastner rows in Annexure E, got {data_rows}"
        )
        assert set(emitted_ids) == expected_ids, (
            f"emitted_ids {set(emitted_ids)} != expected {expected_ids}"
        )

    def test_pof_pitting_with_wide_geometry_excluded(self):
        """The v0.3.3-iteration-1 regression guard: a PITTING feature
        with W > L (incidental geometric circ-ness) must NOT be picked
        by the Annexure E writer. The POF enum is authoritative."""
        from openpyxl import Workbook
        from src.reports.topic_registry import _topic_estimated_erf_circ

        results, _ = self._build_results()
        project = self._build_project()
        wb = Workbook()
        wb.remove(wb.active)

        _topic_estimated_erf_circ(
            wb, "Annexure E", project, results, run2_year=2023,
        )
        ws = wb["Annexure E"]
        emitted_ids = set()
        for row_idx in range(5, ws.max_row + 1):
            fid = ws.cell(row_idx, 2).value
            if fid is not None:
                emitted_ids.add(str(fid))
        # The 5 PITTING features with W > L (anomaly_ids 401, 403,
        # 405, 407, 409 — even index in the test data) must not appear.
        pitting_with_wide_geom = {"401", "403", "405", "407", "409"}
        assert not (emitted_ids & pitting_with_wide_geom), (
            f"PITTING features with W>L leaked into Kastner sheet: "
            f"{emitted_ids & pitting_with_wide_geom}"
        )

    def test_empty_cgrs_produces_no_data_rows(self):
        """No CGR data → header-only sheet, no crash."""
        from openpyxl import Workbook
        from types import SimpleNamespace
        from src.reports.topic_registry import _topic_estimated_erf_circ

        empty = SimpleNamespace(
            cgr_results=[], ffp_results=[],
            repair_predictions=[], flag_report=None,
        )
        project = self._build_project()
        wb = Workbook()
        wb.remove(wb.active)
        _topic_estimated_erf_circ(
            wb, "Annexure E", project, empty, run2_year=2023,
        )
        ws = wb["Annexure E"]
        # Should have header rows but no data rows.
        data_rows = sum(
            1 for r in range(5, ws.max_row + 1)
            if ws.cell(r, 1).value is not None
        )
        assert data_rows == 0
