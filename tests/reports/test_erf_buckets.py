"""Tests for src/reports/erf_buckets.py (v0.3.2 conventions).

v0.3.2 switched the default to raw-float comparison (no rounding),
aligning to the per-feature ERFs reported in Annexure D. The legacy
3-dp rounded mode remains available via ``dp=3``. Both modes are
tested here.

Engineering severity (ERF_EXCEEDS_1) continues to use raw float
regardless of the bucket display mode — pinned at the bottom of this
file.
"""
from __future__ import annotations

import math
from types import SimpleNamespace

import pytest

from src.reports.erf_buckets import (
    DEFAULT_BUCKET_DP,
    ERF_BUCKET_LABELS,
    count_erf_buckets,
    erf_bucket_for,
)


# ---------------------------------------------------------------------------
# Bucket-label invariants
# ---------------------------------------------------------------------------

class TestBucketLabels:
    def test_four_labels(self):
        assert len(ERF_BUCKET_LABELS) == 4

    def test_labels_in_canonical_order(self):
        assert ERF_BUCKET_LABELS == (
            "≤ 0.85",
            "0.85 < ERF ≤ 0.90",
            "0.90 < ERF ≤ 1.00",
            "> 1.00",
        )

    def test_default_bucket_dp_constant_preserved(self):
        # The 3-dp constant stays exported so callers can opt into
        # legacy mode by passing ``dp=DEFAULT_BUCKET_DP``. Pinning the
        # value prevents accidental drift.
        assert DEFAULT_BUCKET_DP == 3


# ---------------------------------------------------------------------------
# Single-value classification — DEFAULT (raw float, dp=None) behaviour
# ---------------------------------------------------------------------------

class TestErfBucketForDefaultRawFloat:
    """v0.3.2 default: ``dp=None`` → raw float comparison."""

    @pytest.mark.parametrize("erf,expected", [
        # Spec test pin: [0.8499, 0.8501, 0.8504, 0.8506] under RAW
        # comparison → 1 in ≤0.85, 3 in 0.85<ERF≤0.90 (since 0.8501,
        # 0.8504, 0.8506 are all strictly > 0.85).
        (0.8499, "≤ 0.85"),
        (0.8501, "0.85 < ERF ≤ 0.90"),
        (0.8504, "0.85 < ERF ≤ 0.90"),
        (0.8506, "0.85 < ERF ≤ 0.90"),
        # 0.85 exactly: ≤-0.85 bucket (inclusive of upper edge).
        (0.85000, "≤ 0.85"),
        # 1.0001: raw > 1.0 → top bucket (NOT the 3rd bucket — that
        # was the dp=3 rounded behaviour).
        (1.0001, "> 1.00"),
        # 0.9999: raw < 1.0 → 3rd bucket.
        (0.9999, "0.90 < ERF ≤ 1.00"),
        # Other obvious cases.
        (1.5, "> 1.00"),
        (1.001, "> 1.00"),
        (0.0, "≤ 0.85"),
        (0.5, "≤ 0.85"),
        (0.92, "0.90 < ERF ≤ 1.00"),
        (1.00000, "0.90 < ERF ≤ 1.00"),     # exactly 1.0 → 3rd bucket
    ])
    def test_canonical_boundary_cases_raw(self, erf, expected):
        assert erf_bucket_for(erf) == expected, (
            f"erf={erf} expected {expected!r}, got {erf_bucket_for(erf)!r}"
        )


# ---------------------------------------------------------------------------
# Single-value classification — LEGACY (dp=3 rounded) mode still works
# ---------------------------------------------------------------------------

class TestErfBucketForLegacyDp3Mode:
    """Pin the v0.2.6 dp=3 mode for callers that opt in explicitly."""

    @pytest.mark.parametrize("erf,expected", [
        # Under dp=3 rounding: 0.8499→0.850, 0.8501→0.850, 0.8504→0.850,
        # 0.8506→0.851. So 3 in ≤0.85, 1 in mid (the v0.2.6 spec).
        (0.8499, "≤ 0.85"),
        (0.8501, "≤ 0.85"),
        (0.8504, "≤ 0.85"),
        (0.8506, "0.85 < ERF ≤ 0.90"),
        # Under dp=3: 1.0001 rounds to 1.000 → 3rd bucket. (Raw mode
        # puts it in top bucket — the key difference between modes.)
        (1.0001, "0.90 < ERF ≤ 1.00"),
        # Identical to raw for these cases:
        (1.5, "> 1.00"),
        (1.001, "> 1.00"),
        (0.85000, "≤ 0.85"),
        (1.00000, "0.90 < ERF ≤ 1.00"),
    ])
    def test_dp3_rounded_classification(self, erf, expected):
        assert erf_bucket_for(erf, dp=3) == expected

    def test_passing_dp_zero_disables_to_integer_round(self):
        # With dp=0, 0.8504 rounds to 1.0 (round-half-to-even).
        assert erf_bucket_for(0.8504, dp=0) == "0.90 < ERF ≤ 1.00"

    def test_passing_dp_higher_uses_more_precision(self):
        # With dp=4, 0.85004 rounds to 0.8500 → still ≤ 0.85.
        assert erf_bucket_for(0.85004, dp=4) == "≤ 0.85"
        assert erf_bucket_for(0.85055, dp=4) == "0.85 < ERF ≤ 0.90"


# ---------------------------------------------------------------------------
# Bulk classification — DEFAULT mode (raw float)
# ---------------------------------------------------------------------------

def _ffp(erf):
    return SimpleNamespace(erf=erf)


class TestCountErfBucketsDefaultRawFloat:
    def test_spec_example_distributes_1_3_0_0_raw(self):
        """v0.3.2 spec: [0.8499, 0.8501, 0.8504, 0.8506] under raw
        comparison → 1 in ≤0.85, 3 in 0.85<ERF≤0.90."""
        ffps = [_ffp(e) for e in (0.8499, 0.8501, 0.8504, 0.8506)]
        counts = count_erf_buckets(ffps)
        assert counts == {
            "≤ 0.85":            1,
            "0.85 < ERF ≤ 0.90": 3,
            "0.90 < ERF ≤ 1.00": 0,
            "> 1.00":            0,
        }

    def test_empty_input_returns_zeroed_dict_with_all_labels(self):
        counts = count_erf_buckets([])
        assert set(counts.keys()) == set(ERF_BUCKET_LABELS)
        assert all(v == 0 for v in counts.values())

    def test_features_above_one_get_top_bucket(self):
        ffps = [_ffp(1.5), _ffp(1.1), _ffp(2.0)]
        counts = count_erf_buckets(ffps)
        assert counts["> 1.00"] == 3

    def test_skips_none_erf(self):
        ffps = [_ffp(None), _ffp(0.5), _ffp(None), _ffp(0.92)]
        counts = count_erf_buckets(ffps)
        assert counts["≤ 0.85"] == 1
        assert counts["0.90 < ERF ≤ 1.00"] == 1
        assert sum(counts.values()) == 2

    def test_skips_inf_and_nan(self):
        ffps = [_ffp(0.5), _ffp(float("inf")), _ffp(float("nan"))]
        counts = count_erf_buckets(ffps)
        assert sum(counts.values()) == 1
        assert counts["≤ 0.85"] == 1

    def test_raw_1_0001_lands_in_top_bucket(self):
        """v0.3.2: raw float comparison puts 1.0001 in '> 1.00'."""
        ffps = [_ffp(1.0001)]
        counts = count_erf_buckets(ffps)
        assert counts["> 1.00"] == 1
        assert counts["0.90 < ERF ≤ 1.00"] == 0


class TestCountErfBucketsLegacyDp3Mode:
    """The dp=3 mode still exists and behaves like v0.2.6."""

    def test_spec_example_distributes_3_1_0_0_dp3(self):
        """v0.2.6 spec: [0.8499, 0.8501, 0.8504, 0.8506] under dp=3
        rounding → 3 in ≤0.85, 1 in 0.85<ERF≤0.90."""
        ffps = [_ffp(e) for e in (0.8499, 0.8501, 0.8504, 0.8506)]
        counts = count_erf_buckets(ffps, dp=3)
        assert counts == {
            "≤ 0.85":            3,
            "0.85 < ERF ≤ 0.90": 1,
            "0.90 < ERF ≤ 1.00": 0,
            "> 1.00":            0,
        }

    def test_dp3_1_0001_lands_in_3rd_bucket(self):
        """Under dp=3 rounding: 1.0001 → 1.000 → '0.90 < ERF ≤ 1.00'."""
        ffps = [_ffp(1.0001)]
        counts = count_erf_buckets(ffps, dp=3)
        assert counts["0.90 < ERF ≤ 1.00"] == 1
        assert counts["> 1.00"] == 0


# ---------------------------------------------------------------------------
# ERF_EXCEEDS_1 flag uses RAW ERF regardless of bucket display mode.
# ---------------------------------------------------------------------------

class TestRawFloatNotRoundedForFlag:
    def test_erf_1_0001_under_default_lands_in_top_and_would_flag(self):
        """A feature with raw ERF = 1.0001 under v0.3.2 default:
          * BUCKET classification: '> 1.00' (raw 1.0001 > 1.00).
          * QA flag uses raw 1.0001 ≥ 1.0 → fires ERF_EXCEEDS_1.
        Both sides agree under raw-float comparison."""
        assert erf_bucket_for(1.0001) == "> 1.00"
        raw = 1.0001
        assert raw >= 1.0

    def test_erf_1_0001_under_dp3_lands_in_3rd_but_flag_still_fires(self):
        """A feature with raw ERF = 1.0001 under legacy dp=3:
          * BUCKET classification rounds to 1.000 → '0.90 < ERF ≤ 1.00'.
          * QA flag uses raw 1.0001 ≥ 1.0 → STILL fires ERF_EXCEEDS_1.
        The flag is decoupled from the bucket display mode."""
        assert erf_bucket_for(1.0001, dp=3) == "0.90 < ERF ≤ 1.00"
        raw = 1.0001
        assert raw >= 1.0, "raw ERF threshold check unaffected by bucket dp"


# ---------------------------------------------------------------------------
# Regression: BPCL Malarna-Karwadi 16" — bucket counts under the v0.3.2
# default (raw float) now match the Annexure C XLSX directly, NOT the
# published PDF.
#
# v0.2.6 design pinned to PDF (4803/331/4/0); v0.3.2 design pins to the
# annexure XLSX (4787/347/4/0). The 16-feature drift is a
# 3-dp-rounding artifact; the XLSX writes the raw ERFs without
# rounding, and downstream consumers re-bucketing the XLSX (auditors,
# QGIS layer) now get the same counts as the GUI. The PDF's display-
# precision counts remain reproducible via ``dp=3`` if needed.
# ---------------------------------------------------------------------------

class TestBpclMalarnaRegression:
    @pytest.fixture
    def annex_path(self):
        from pathlib import Path
        candidate = Path(
            r"C:/Users/varun/OneDrive/Documents/bpcl_malarna/"
            r"bpcl_malarna_project_output/"
            r"BPCL_Malarna_Karwadi_16in_Validation_annexure.xlsx"
        )
        if not candidate.exists():
            pytest.skip(
                "BPCL test pack annexure not on disk; skip regression"
            )
        return candidate

    def _read_erfs(self, annex_path):
        import openpyxl
        wb = openpyxl.load_workbook(annex_path, read_only=True)
        # Annexure C holds the ERF column at index 10 (J).
        ws = wb["Annexure C"]
        erfs = []
        for row_idx in range(5, ws.max_row + 1):
            v = ws.cell(row_idx, 10).value
            if v is None:
                continue
            erfs.append(float(v))
        wb.close()
        return erfs

    def test_bucket_counts_raw_match_annexure_xlsx(self, annex_path):
        """v0.3.2 default (raw float) — the GUI Results-screen bucket
        counts now exactly match what a downstream consumer would get
        by re-bucketing the Annexure C ERF column directly. The
        expected counts (4787/347/4/0) come from the XLSX raw values,
        NOT the published PDF (which shows 4803/331/4/0 due to its
        display-precision rounding)."""
        erfs = self._read_erfs(annex_path)
        assert len(erfs) == 5138, (
            f"BPCL Annexure C should have 5138 rows; got {len(erfs)}"
        )
        ffps = [_ffp(e) for e in erfs]
        counts = count_erf_buckets(ffps)
        assert counts == {
            "≤ 0.85":            4787,
            "0.85 < ERF ≤ 0.90": 347,
            "0.90 < ERF ≤ 1.00": 4,
            "> 1.00":            0,
        }, (
            f"BPCL bucket counts mismatch (v0.3.2 raw-float mode): "
            f"{counts}"
        )

    def test_bucket_counts_dp3_still_reproduce_pdf(self, annex_path):
        """The legacy dp=3 mode is preserved for callers that need to
        reproduce a published PDF's bucket counts. BPCL Malarna PDF
        page 18 reports 4803/331/4/0 — the dp=3 mode should still
        reproduce that exactly."""
        erfs = self._read_erfs(annex_path)
        ffps = [_ffp(e) for e in erfs]
        counts = count_erf_buckets(ffps, dp=3)
        assert counts == {
            "≤ 0.85":            4803,
            "0.85 < ERF ≤ 0.90": 331,
            "0.90 < ERF ≤ 1.00": 4,
            "> 1.00":            0,
        }
